from flask import Flask, render_template, request, jsonify, send_file, abort, redirect, url_for, make_response
import sqlite3, os, io, uuid, socket, time, json, secrets, urllib.request
from urllib.parse import quote
from datetime import date, datetime, timedelta
from functools import wraps
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import qrcode, bcrypt, pyotp
import jwt as pyjwt

# ── Load .env if exists ────────────────────────────────────────────────────────
def _load_env():
    env_path = os.path.join(os.path.dirname(__file__), '.env')
    if os.path.exists(env_path):
        with open(env_path, encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    k, v = line.split('=', 1)
                    os.environ.setdefault(k.strip(), v.strip())
_load_env()

app = Flask(__name__)
app.config['TEMPLATES_AUTO_RELOAD'] = True

# ── Security Config ─────────────────────────────────────────────────────────────
_secret = os.environ.get('SECRET_KEY', '')
if not _secret or _secret == 'CHANGE_ME_USE_RANDOM_64_CHAR_STRING_IN_PRODUCTION':
    _secret = secrets.token_hex(32)
    print('  [!] SECRET_KEY не задан в .env — используется временный ключ.')
    print('  [!] Создайте файл .env и задайте SECRET_KEY для продакшна!')
app.config['SECRET_KEY'] = _secret

JWT_EXPIRY         = int(os.environ.get('JWT_EXPIRY', 60 * 60 * 12))   # 12h
SECURE_COOKIES     = os.environ.get('SECURE_COOKIES', 'False').lower() == 'true'
MAX_UPLOAD_MB      = int(os.environ.get('MAX_UPLOAD_MB', '16'))
app.config['MAX_CONTENT_LENGTH'] = MAX_UPLOAD_MB * 1024 * 1024

ALLOWED_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.heic'}

app.jinja_env.filters['from_json'] = json.loads

DB      = os.path.join(os.path.dirname(__file__), "inventory.db")
UPLOADS = os.path.join(os.path.dirname(__file__), "static", "photos")
os.makedirs(UPLOADS, exist_ok=True)
SIGS = os.path.join(os.path.dirname(__file__), "static", "signatures")
os.makedirs(SIGS, exist_ok=True)

CATEGORIES = ["Ноутбук","Монитор","Кресло","Стол","Клавиатура","Мышь","Принтер","Телефон","Другое"]
PREFIXES   = {"Ноутбук":"НТБ","Монитор":"МОН","Кресло":"КРС","Стол":"СТЛ",
               "Клавиатура":"КЛВ","Мышь":"МШ","Принтер":"ПРН","Телефон":"ТЛФ","Другое":"ДРГ"}
CONDITIONS = ["Хорошее","Потёрто","Требует ремонта","Списано","Утеряно"]
STATUSES   = ["Занято","Свободно"]

ROLES = {
    "superadmin": {"label":"Супер-Админ",   "color":"5856D6","can_manage_users":True, "can_delete":True, "can_edit":True, "can_view_all":True, "can_issue":True, "can_export":True, "can_approve":True},
    "aho":        {"label":"АХО / IT",      "color":"007AFF","can_manage_users":False,"can_delete":False,"can_edit":True, "can_view_all":True, "can_issue":True, "can_export":True, "can_approve":True},
    "hr":         {"label":"HR",            "color":"34C759","can_manage_users":False,"can_delete":False,"can_edit":False,"can_view_all":True, "can_issue":True, "can_export":False,"can_approve":False},
    "employee":   {"label":"Сотрудник",     "color":"8E8E93","can_manage_users":False,"can_delete":False,"can_edit":False,"can_view_all":False,"can_issue":False,"can_export":False,"can_approve":False},
    "auditor":    {"label":"Аудитор",       "color":"5AC8FA","can_manage_users":False,"can_delete":False,"can_edit":False,"can_view_all":True, "can_issue":False,"can_export":True, "can_approve":False},
    "deputy":     {"label":"Зам. Директора","color":"FF9500","can_manage_users":False,"can_delete":False,"can_edit":False,"can_view_all":True, "can_issue":False,"can_export":True, "can_approve":True},
    "director":   {"label":"Ген. Директор", "color":"FF3B30","can_manage_users":True, "can_delete":True, "can_edit":True, "can_view_all":True, "can_issue":True, "can_export":True, "can_approve":True},
    "accountant": {"label":"Бухгалтер",     "color":"30B0C7","can_manage_users":False,"can_delete":False,"can_edit":False,"can_view_all":True, "can_issue":False,"can_export":True, "can_approve":True},
    "viewer":     {"label":"Наблюдатель",   "color":"5AC8FA","can_manage_users":False,"can_delete":False,"can_edit":False,"can_view_all":True, "can_issue":False,"can_export":False,"can_approve":False},
}

# ─── DATABASE ────────────────────────────────────────────────────────────────
def get_db():
    db_file = os.path.join(os.path.dirname(__file__), "inventory.db")
    # Increase timeout to 20 seconds to prevent "database is locked" during heavy concurrent access
    conn = sqlite3.connect(db_file, timeout=20.0)
    
    # Enable WAL (Write-Ahead Logging) for high concurrency (Production Readiness)
    conn.execute('PRAGMA journal_mode=WAL;')
    conn.execute('PRAGMA synchronous=NORMAL;')
    conn.execute('PRAGMA cache_size=-64000;') # 64MB cache
    
    conn.row_factory = sqlite3.Row
    return conn

def add_col(db, table, col, t):
    try:
        db.execute(f"ALTER TABLE {table} ADD COLUMN {col} {t}")
    except sqlite3.OperationalError as e:
        if "duplicate column name" not in str(e).lower():
            print(f"  [!] Ошибка миграции ({table}.{col}): {e}")
    except Exception as e:
        print(f"  [!] Ошибка при добавлении колонки {col}: {e}")

def migrate_db():
    """Safe migrations for existing databases."""
    with get_db() as db:
        migrations = [
            ("dismissals","deadline","DATE"),
            ("dismissals","item_conditions","TEXT DEFAULT '{}'"),
            ("dismissals","item_comments","TEXT DEFAULT '{}'"),
            ("dismissals","confirmed_signature","INTEGER DEFAULT 0"),
            ("users", "department", "TEXT"),
            ("users", "token_version", "INTEGER DEFAULT 0"),
            ("users", "force_password_change", "INTEGER DEFAULT 0"),
            ("users", "totp_secret", "TEXT"),
            ("users", "totp_enabled", "INTEGER DEFAULT 0"),
            ("users", "telegram_chat_id", "TEXT"),
            ("users", "expires_at", "DATE"),
            ("users", "onboarding_done", "INTEGER DEFAULT 1"),
            ("users", "last_login", "TIMESTAMP"),
            ("users", "avatar_color", "TEXT"),
            ("users", "created_at", "TIMESTAMP DEFAULT CURRENT_TIMESTAMP"),
            ("items", "purchase_price", "REAL"),
            ("items", "purchase_date", "DATE"),
            ("items", "supplier", "TEXT"),
            ("items", "warranty_until", "DATE"),
            ("items", "check_date", "TEXT"),
            ("items", "employee_id", "INTEGER"),
            ("issuances", "signature", "TEXT"),
            ("returns", "signature", "TEXT"),
            ("history", "field", "TEXT"),
            ("history", "new_val", "TEXT"),
            ("documents", "employee_id", "INTEGER"),
            ("documents", "employee_name", "TEXT"),
            ("dismissals", "signature", "TEXT"),
            ("dismissals", "aho_signature", "TEXT"),
            ("dismissals", "it_signature", "TEXT"),
            ("dismissals", "hr_signature", "TEXT"),
            ("dismissals", "hr_at", "TIMESTAMP"),
            ("dismissals", "hr_by_id", "INTEGER"),
            ("dismissals", "hr_by_name", "TEXT"),
            ("dismissals", "employee_signature", "TEXT"),
            ("documents", "signature", "TEXT"),
            ("doc_approvals", "signature", "TEXT"),
            ("documents", "employee_id", "INTEGER"),
            ("documents", "employee_name", "TEXT"),
        ]
        for table, col, typ in migrations:
            add_col(db, table, col, typ)

def init_db():
    with get_db() as db:
        # ─── Core tables ───
        db.execute("""CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL, email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL, role TEXT NOT NULL DEFAULT 'employee',
            department TEXT, active INTEGER DEFAULT 1,
            token_version INTEGER DEFAULT 0,
            force_password_change INTEGER DEFAULT 0,
            totp_secret TEXT, totp_enabled INTEGER DEFAULT 0,
            telegram_chat_id TEXT, expires_at DATE,
            onboarding_done INTEGER DEFAULT 0,
            last_login TIMESTAMP, avatar_color TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
        db.execute("""CREATE TABLE IF NOT EXISTS items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            place TEXT NOT NULL, inv_num TEXT NOT NULL UNIQUE, category TEXT NOT NULL,
            model TEXT, serial_num TEXT, room TEXT NOT NULL,
            employee TEXT, employee_id INTEGER,
            status TEXT DEFAULT 'Свободно', condition TEXT DEFAULT 'Хорошее',
            check_date TEXT, notes TEXT, photo TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
        db.execute("""CREATE TABLE IF NOT EXISTS history (
            id INTEGER PRIMARY KEY AUTOINCREMENT, item_id INTEGER NOT NULL,
            user_id INTEGER, user_name TEXT, action TEXT NOT NULL,
            field TEXT, old_val TEXT, new_val TEXT,
            ts TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
        db.execute("""CREATE TABLE IF NOT EXISTS issuances (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id INTEGER NOT NULL, employee_name TEXT NOT NULL,
            issued_by INTEGER NOT NULL, issued_by_name TEXT NOT NULL,
            items_json TEXT NOT NULL, status TEXT DEFAULT 'pending',
            signature TEXT,
            confirmed_at TIMESTAMP, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
        db.execute("""CREATE TABLE IF NOT EXISTS returns (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id INTEGER NOT NULL, employee_name TEXT NOT NULL,
            initiated_by INTEGER NOT NULL, initiated_by_name TEXT NOT NULL,
            items_json TEXT NOT NULL, photos_json TEXT,
            accepted_by INTEGER, accepted_by_name TEXT,
            status TEXT DEFAULT 'pending', signature TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            completed_at TIMESTAMP)""")
        db.execute("""CREATE TABLE IF NOT EXISTS dismissals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id INTEGER NOT NULL, employee_name TEXT NOT NULL,
            employee_email TEXT,
            initiated_by INTEGER NOT NULL, initiated_by_name TEXT NOT NULL,
            items_json TEXT NOT NULL,
            status TEXT DEFAULT 'pending',
            notes TEXT,
            photos_json TEXT, signature TEXT,
            confirmed_by INTEGER, confirmed_by_name TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            completed_at TIMESTAMP)""")
        db.execute("""CREATE TABLE IF NOT EXISTS login_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER, email TEXT, success INTEGER,
            ip TEXT, user_agent TEXT,
            ts TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
        db.execute("""CREATE TABLE IF NOT EXISTS maintenance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_id INTEGER NOT NULL,
            reported_by_id INTEGER,
            reported_by_name TEXT,
            description TEXT,
            priority TEXT DEFAULT 'medium',
            status TEXT DEFAULT 'pending',
            resolved_by TEXT,
            resolved_at TIMESTAMP,
            resolution TEXT,
            rejection_reason TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
        db.execute("""CREATE TABLE IF NOT EXISTS asset_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id INTEGER NOT NULL,
            employee_name TEXT NOT NULL,
            category TEXT NOT NULL,
            reason TEXT,
            status TEXT DEFAULT 'pending',
            rejection_reason TEXT,
            resolved_by TEXT,
            resolved_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
        db.execute("""CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            signed_by_id INTEGER,
            signed_by_name TEXT,
            item_count INTEGER DEFAULT 0,
            note TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
        db.execute("""CREATE TABLE IF NOT EXISTS revoked_tokens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            token_version INTEGER NOT NULL,
            revoked_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
        db.execute("""CREATE TABLE IF NOT EXISTS api_keys (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            key_hash TEXT UNIQUE NOT NULL,
            scopes TEXT DEFAULT 'read',
            user_id INTEGER,
            last_used TIMESTAMP,
            expires_at DATE,
            active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
        db.execute("""CREATE TABLE IF NOT EXISTS inventory_sessions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            status TEXT DEFAULT 'active',
            created_by_id INTEGER,
            created_by_name TEXT,
            department TEXT,
            total_items INTEGER DEFAULT 0,
            checked_items INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            completed_at TIMESTAMP)""")
        db.execute("""CREATE TABLE IF NOT EXISTS inventory_checks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id INTEGER NOT NULL,
            item_id INTEGER NOT NULL,
            status TEXT DEFAULT 'pending',
            checked_by_id INTEGER,
            checked_by_name TEXT,
            photo TEXT,
            note TEXT,
            checked_at TIMESTAMP,
            FOREIGN KEY (session_id) REFERENCES inventory_sessions(id))""")
        db.execute("""CREATE TABLE IF NOT EXISTS equipment_templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            description TEXT,
            items_json TEXT DEFAULT '[]',
            created_by TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
        
        # Migrations / Column updates
        add_col(db, "items", "employee_id", "INTEGER")
        add_col(db, "items", "photo", "TEXT")
        add_col(db, "history", "user_id", "INTEGER")
        add_col(db, "history", "user_name", "TEXT")
        # ─── Docflow tables ───
        _init_docflow_tables(db)
        if db.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
            pw = bcrypt.hashpw(b"admin123", bcrypt.gensalt()).decode()
            db.execute("INSERT INTO users (name,email,password_hash,role) VALUES (?,?,?,?)",
                       ("Администратор","admin@tracko.uz",pw,"superadmin"))
            print("  👤  admin@tracko.uz / admin123")

def log_h(db, item_id, action, field=None, old_val=None, new_val=None, uid=None, uname=None):
    db.execute("INSERT INTO history (item_id,user_id,user_name,action,field,old_val,new_val) VALUES (?,?,?,?,?,?,?)",
               (item_id,uid,uname,action,field,old_val,new_val))

# ─── AUTH ─────────────────────────────────────────────────────────────────────
_login_attempts = {}  # ip -> {count, locked_until}

def check_rate_limit(ip):
    """Returns (allowed, seconds_left)"""
    now = time.time()
    data = _login_attempts.get(ip, {"count": 0, "locked_until": 0})
    if now < data["locked_until"]:
        return False, int(data["locked_until"] - now)
    return True, 0

def record_failed_login(ip):
    now = time.time()
    data = _login_attempts.get(ip, {"count": 0, "locked_until": 0})
    data["count"] = data.get("count", 0) + 1
    if data["count"] >= 5:
        data["locked_until"] = now + 300  # 5 min
        data["count"] = 0
    _login_attempts[ip] = data

def clear_failed_login(ip):
    _login_attempts.pop(ip, None)
def make_token(u):
    return pyjwt.encode({"sub":u["id"],"role":u["role"],"name":u["name"],
                          "tv": u.get("token_version",0),
                          "exp":int(time.time())+JWT_EXPIRY},
                        app.config["SECRET_KEY"], algorithm="HS256")

def get_current_user():
    token = request.cookies.get("token") or request.headers.get("Authorization","").replace("Bearer ","")
    if not token: return None
    try:
        p = pyjwt.decode(token, app.config["SECRET_KEY"], algorithms=["HS256"])
        with get_db() as db:
            u = db.execute("SELECT * FROM users WHERE id=? AND active=1",(p["sub"],)).fetchone()
        if not u: return None
        u = dict(u)
        # Check token_version (invalidated on logout/password change)
        token_ver = p.get("tv", 0)
        if token_ver != (u.get("token_version") or 0):
            return None
        # Check account expiry
        if u.get("expires_at"):
            from datetime import date as _date
            if _date.today().isoformat() > u["expires_at"]:
                return None
        return u
    except Exception as e:
        app.logger.error(f"Save photo error: {e}")
        return None

def login_required(f):
    @wraps(f)
    def dec(*a,**kw):
        u = get_current_user()
        if not u:
            if request.is_json: return jsonify({"error":"Не авторизован"}),401
            return redirect("/login")
        request.current_user = u
        return f(*a,**kw)
    return dec

def roles_required(*roles):
    def decorator(f):
        @wraps(f)
        @login_required
        def dec(*a,**kw):
            if request.current_user["role"] not in roles:
                if request.is_json: return jsonify({"error":"Нет доступа"}),403
                abort(403)
            return f(*a,**kw)
        return dec
    return decorator

def get_lan_ip():
    try:
        s=socket.socket(socket.AF_INET,socket.SOCK_DGRAM); s.connect(("8.8.8.8",80))
        ip=s.getsockname()[0]; s.close(); return ip
    except: return "127.0.0.1"

def next_inv(cat):
    p=PREFIXES.get(cat,"ДРГ")
    with get_db() as db:
        rows=db.execute("SELECT inv_num FROM items WHERE inv_num LIKE ?",(f"{p}-%",)).fetchall()
    nums=[int(r["inv_num"].split("-")[1]) for r in rows if r["inv_num"].split("-")[1].isdigit()]
    return f"{p}-{(max(nums)+1 if nums else 1):03d}"

def _save_signature(base64_str, prefix):
    if not base64_str or "," not in base64_str: return None
    import base64
    try:
        header, encoded = base64_str.split(",", 1)
        data = base64.b64decode(encoded)
        fname = f"{prefix}_{uuid.uuid4().hex[:8]}.png"
        fpath = os.path.join(SIGS, fname)
        with open(fpath, "wb") as f: f.write(data)
        return "/static/signatures/" + fname
    except Exception as e:
        app.logger.error(f"Save signature error: {e}")
        return None

def qr_png(url):
    qr=qrcode.QRCode(version=1,error_correction=qrcode.constants.ERROR_CORRECT_M,box_size=8,border=2)
    qr.add_data(url); qr.make(fit=True)
    buf=io.BytesIO(); qr.make_image(fill_color="black",back_color="white").save(buf,"PNG"); buf.seek(0)
    return buf

def bhost():
    # Priority: Env Var > X-Forwarded-Host > Request Host
    base = os.environ.get('BASE_URL')
    if base: return base.rstrip("/")
    
    # Check for proxy headers first (important for Docker/Gunicorn)
    forwarded = request.headers.get("X-Forwarded-Host")
    if forwarded:
        proto = request.headers.get("X-Forwarded-Proto", "http")
        return f"{proto}://{forwarded}"
        
    h = request.host_url.rstrip("/")
    # If on localhost inside container, try to get reachable IP
    if "localhost" in h or "127.0.0.1" in h:
        lan = get_lan_ip()
        return h.replace("localhost", lan).replace("127.0.0.1", lan)
    return h

@app.route('/force-admin')
def force_admin():
    res = make_response(redirect('/login'))
    res.set_cookie('token', '', expires=0)
    return res

# ── Security Headers ───────────────────────────────────────────────────────────
@app.after_request
def add_security_headers(resp):
    resp.headers['X-Content-Type-Options'] = 'nosniff'
    resp.headers['X-Frame-Options'] = 'SAMEORIGIN'
    resp.headers['X-Content-Type-Options'] = 'nosniff'
    resp.headers['X-XSS-Protection'] = '1; mode=block'
    resp.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    resp.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://unpkg.com https://cdnjs.cloudflare.com; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "img-src 'self' data: blob:; "
        "connect-src 'self';"
    )
    resp.headers['X-XSS-Protection'] = '1; mode=block'
    resp.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    resp.headers['Permissions-Policy'] = 'geolocation=(), microphone=(), payment=()'
    # Cache control for API
    if request.path.startswith('/api/'):
        resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, private'
        resp.headers['Pragma'] = 'no-cache'
    # Remove server version info
    resp.headers.pop('Server', None)
    return resp

# ── Global error handlers ──────────────────────────────────────────────────────
@app.errorhandler(404)
def not_found(e):
    if request.is_json or request.path.startswith('/api/'):
        return jsonify({'error': 'Не найдено'}), 404
    u = get_current_user()
    if u: return redirect('/dashboard')
    return render_template('login.html'), 404

@app.errorhandler(403)
def forbidden(e):
    if request.is_json or request.path.startswith('/api/'):
        return jsonify({'error': 'Нет доступа'}), 403
    return redirect('/login')

@app.errorhandler(413)
def too_large(e):
    return jsonify({'error': f'Файл слишком большой. Максимум {MAX_UPLOAD_MB}MB'}), 413

@app.errorhandler(500)
def server_error(e):
    import traceback
    traceback.print_exc()
    app.logger.error(f'500 error: {e}')
    if request.is_json or request.path.startswith('/api/'):
        return jsonify({'error': 'Внутренняя ошибка сервера'}), 500
    return redirect("/dashboard")


# ─── AUTH ROUTES ──────────────────────────────────────────────────────────────
@app.route("/login")
def login_page():
    if get_current_user(): return redirect("/dashboard")
    return render_template("login.html")

@app.route("/api/auth/login", methods=["POST"])
def api_login():
    d = request.json or {}
    ip = request.remote_addr or "unknown"
    ua = request.headers.get("User-Agent", "")[:200]
    # Rate limit check
    allowed, secs = check_rate_limit(ip)
    if not allowed:
        return jsonify({"error": f"Слишком много попыток. Подождите {secs} сек."}), 429
    with get_db() as db:
        u = db.execute("SELECT * FROM users WHERE email=? AND active=1", (d.get("email","").lower(),)).fetchone()
    ok = u and bcrypt.checkpw(d.get("password","").encode(), u["password_hash"].encode())
    with get_db() as db:
        db.execute("INSERT INTO login_log (user_id,email,success,ip,user_agent) VALUES (?,?,?,?,?)",
                   (u["id"] if u else None, d.get("email","").lower(), 1 if ok else 0, ip, ua))
    if not ok:
        record_failed_login(ip)
        return jsonify({"error": "Неверный email или пароль"}), 401
    clear_failed_login(ip)
    # Update last_login timestamp
    with get_db() as db:
        db.execute("UPDATE users SET last_login=CURRENT_TIMESTAMP WHERE id=?", (u["id"],))
    u_dict = dict(u)
    result = {
        "ok": True,
        "role": u["role"],
        "name": u["name"],
        "force_password_change": bool(u_dict.get("force_password_change", 0)),
        "onboarding_done": bool(u_dict.get("onboarding_done", 1)),
    }
    resp = jsonify(result)
    resp.set_cookie(
        "token", make_token(u_dict),
        httponly=True,
        samesite="Lax",
        secure=SECURE_COOKIES,
        max_age=JWT_EXPIRY
    )
    return resp

@app.route("/api/auth/logout",methods=["POST"])
@login_required
def api_logout():
    u = request.current_user
    with get_db() as db:
        db.execute("UPDATE users SET token_version=COALESCE(token_version,0)+1 WHERE id=?", (u["id"],))
    resp = make_response(jsonify({"ok": True}))
    resp.delete_cookie("token")
    return resp

@app.route("/api/auth/me")
@login_required
def api_me():
    u=dict(request.current_user); u.pop("password_hash",None)
    u["role_info"]=ROLES.get(u["role"],{})
    return jsonify(u)

# Asset history is handled by api_history and get_item_history below

# ─── PAGES ────────────────────────────────────────────────────────────────────
@app.route("/")
def index_redirect():
    if get_current_user(): return redirect("/dashboard")
    return redirect("/login")

@app.route("/dashboard")
@login_required
def index():
    u=request.current_user
    with get_db() as db:
        all_emps = [r["name"] for r in db.execute("SELECT name FROM users WHERE active=1 ORDER BY name").fetchall()]
    return render_template("index.html",categories=CATEGORIES,conditions=CONDITIONS,
                           statuses=STATUSES,user=u,role_info=ROLES.get(u["role"],{}),roles=ROLES,
                           current_user=u, all_emps=all_emps)

@app.route("/asset/<inv_num>")
def asset_page(inv_num):
    with get_db() as db:
        item=db.execute("SELECT * FROM items WHERE inv_num=?",(inv_num,)).fetchone()
    if not item: abort(404)
    item=dict(item)
    u=get_current_user()
    if u:
        can_edit=ROLES[u["role"]]["can_edit"]
        is_owner=str(item.get("employee_id"))==str(u["id"]) or item.get("employee")==u["name"]
        role=u["role"]
    else:
        can_edit=False
        is_owner=False
        role="guest"

    # Financial Calculations
    fin = {"residual_value": None, "warranty_status": "unknown", "depreciation_pct": 0}
    if item.get("purchase_price") and item.get("purchase_date"):
        try:
            from datetime import date as _d
            purchased = _d.fromisoformat(item["purchase_date"])
            today = _d.today()
            useful_life = {"Ноутбук": 3, "Монитор": 5, "Кресло": 7, "Стол": 10, "Принтер": 4, "Телефон": 2}.get(item.get("category"), 5)
            years_used = (today - purchased).days / 365.25
            fin["depreciation_pct"] = min(100.0, round((years_used / useful_life) * 100, 1))
            fin["residual_value"] = max(0, round(item["purchase_price"] * (1 - fin["depreciation_pct"] / 100), 2))
        except: pass
    
    if item.get("warranty_until"):
        try:
            from datetime import date as _d
            if _d.fromisoformat(item["warranty_until"]) > _d.today(): fin["warranty_status"] = "active"
            else: fin["warranty_status"] = "expired"
        except: pass

    return render_template("asset.html",item=item,user=u,can_edit=can_edit,is_owner=is_owner,role=role, financials=fin)

@app.route("/employee/<path:name>")
@login_required
def employee_page(name):
    with get_db() as db:
        items=db.execute("SELECT * FROM items WHERE employee=? ORDER BY category",(name,)).fetchall()
    return render_template("employee.html",employee=name,items=[dict(i) for i in items],user=request.current_user)

@app.route("/api/employee/<path:name>/export")
@login_required
def export_employee_items(name):
    u = request.current_user
    if u["role"] == "employee" and u["name"] != name:
        return jsonify({"error": "Нет доступа"}), 403
    with get_db() as db:
        rows = db.execute("SELECT * FROM items WHERE employee=? ORDER BY category", (name,)).fetchall()
    wb=Workbook(); ws=wb.active; ws.title=f"Техника {name}"
    headers=["Инв. №","Категория","Модель","Серийный №","Кабинет","Состояние","Примечания"]
    ws.append(headers)
    for row in rows:
        ws.append([row["inv_num"], row["category"], row["model"] or "", row["serial_num"] or "—", row["room"], row["condition"], row["notes"] or ""])
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,as_attachment=True,download_name=f"Items_{name.replace(' ','_')}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/admin/users")
@roles_required("superadmin","aho","director")
def admin_users_page():
    with get_db() as db:
        users=db.execute("SELECT id,name,email,role,department,active FROM users ORDER BY role,name").fetchall()
    return render_template("admin_users.html",users=[dict(u) for u in users],
                           roles=ROLES,user=request.current_user)

@app.route("/qr-sheet")
@login_required
def qr_sheet():
    with get_db() as db:
        items=db.execute("SELECT id,inv_num,category,model,room,employee,place FROM items ORDER BY room,inv_num").fetchall()
    return render_template("qr_sheet.html",items=[dict(i) for i in items])

@app.route("/qr-sheet-employees")
@login_required
def qr_sheet_employees():
    with get_db() as db:
        rows=db.execute("SELECT employee,COUNT(*) as cnt FROM items WHERE employee IS NOT NULL AND employee!='' AND employee!='—' GROUP BY employee ORDER BY employee").fetchall()
    emps=[{"name":r["employee"],"name_encoded":quote(r["employee"]),"count":r["cnt"]} for r in rows]
    return render_template("qr_sheet_employees.html",employees=emps)

@app.route("/qr-print")
@login_required
def qr_print_page():
    """Advanced QR print page: filter by room, category, employee, department. Thermal printer support."""
    with get_db() as db:
        items = db.execute("""
            SELECT i.id, i.inv_num, i.category, i.model, i.room, i.employee,
                   i.place, i.serial_num, i.condition,
                   u.department
            FROM items i
            LEFT JOIN users u ON u.name = i.employee AND i.employee != '—'
            ORDER BY i.room, i.category, i.inv_num
        """).fetchall()
        rooms  = [r[0] for r in db.execute("SELECT DISTINCT room FROM items WHERE room IS NOT NULL ORDER BY room").fetchall()]
        cats   = [r[0] for r in db.execute("SELECT DISTINCT category FROM items ORDER BY category").fetchall()]
        emps   = [r[0] for r in db.execute("SELECT DISTINCT employee FROM items WHERE employee IS NOT NULL AND employee != '—' ORDER BY employee").fetchall()]
        depts  = [r[0] for r in db.execute("SELECT DISTINCT department FROM users WHERE department IS NOT NULL AND department != '' ORDER BY department").fetchall()]
    u = request.current_user
    return render_template("qr_print.html",
        items=[dict(i) for i in items],
        rooms=rooms, categories=cats, employees=emps, departments=depts,
        user=u, role_info=ROLES.get(u["role"], {}), host=bhost())

# ─── QR ───────────────────────────────────────────────────────────────────────
@app.route("/api/user/<int:uid>/qr")
@login_required
def get_user_qr(uid):
    """Generates a QR code for an employee's profile."""
    return send_file(qr_png(f"{bhost()}/user/{uid}"), mimetype="image/png")

@app.route("/api/qr/<path:inv_num>")
@login_required
def get_item_qr(inv_num):
    """Generates a QR code for an item."""
    try:
        url = f"{bhost()}/asset/{inv_num}"
        img_buf = qr_png(url)
        return send_file(img_buf, mimetype="image/png", max_age=0)
    except Exception as e:
        app.logger.error(f"QR Error: {e}")
        return abort(500)

@app.route("/api/qr_employee/<path:enc_name>")
@login_required
def get_employee_qr(enc_name):
    """Generates a QR code for an employee name string."""
    return send_file(qr_png(f"{bhost()}/employee/{enc_name}"), mimetype="image/png")

@app.route("/employee/<int:uid>/print")
@login_required
def print_employee_label(uid):
    """Page for printing an employee's QR badge."""
    with get_db() as db:
        user = db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if not user: abort(404)
    return render_template("print_employee.html", user=dict(user), host=bhost())

@app.route("/asset/<inv_num>/print")
@login_required
def print_label_page(inv_num):
    with get_db() as db:
        item = db.execute("SELECT * FROM items WHERE inv_num=?", (inv_num,)).fetchone()
    if not item: abort(404)
    return render_template("print_label.html", item=dict(item), host=bhost())

# ─── USERS API ────────────────────────────────────────────────────────────────
@app.route("/api/users")
@roles_required("superadmin","aho","hr","director")
def get_users():
    with get_db() as db:
        rows=db.execute("SELECT id,name,email,role,department,active,created_at FROM users ORDER BY name").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/users",methods=["POST"])
@roles_required("superadmin","aho","director")
def create_user():
    d=request.json
    if not d.get("email") or not d.get("name") or not d.get("password"):
        return jsonify({"error":"Заполни все поля"}),400
    pw_raw = d["password"]
    if len(pw_raw) < 8:
        return jsonify({"error":"Пароль минимум 8 символов"}),400
    try:
        pw=bcrypt.hashpw(pw_raw.encode(),bcrypt.gensalt()).decode()
        role = d.get("role", "employee")
        if role not in ROLES:
            for rk, rv in ROLES.items():
                if rv["label"] == role:
                    role = rk; break
            else: role = "employee"

        with get_db() as db:
            cur = db.execute("""INSERT INTO users (name,email,password_hash,role,department,force_password_change)
                          VALUES (?,?,?,?,?,1)""",
                       (d["name"], d["email"].lower(), pw, role, d.get("department", "")))
            new_id = cur.lastrowid
        return jsonify({"ok":True, "id": new_id})
    except sqlite3.IntegrityError:
        return jsonify({"error":"Email уже используется"}),400
    except Exception as e:
        return jsonify({"error":str(e)}),500

@app.route("/api/users/<int:uid>",methods=["PUT"])
@roles_required("superadmin","aho","director")
def update_user(uid):
    d=request.json or {}; u=request.current_user
    if u["role"]=="aho" and d.get("role")=="superadmin":
        return jsonify({"error":"Нет прав"}),403
    ALLOWED = frozenset({"name","email","role","active","department","doc_role",
                         "telegram_chat_id","expires_at","avatar_color"})
    sets=[]; vals=[]
    for k,v in d.items():
        if k in ALLOWED and k != "active":
            sets.append(f"{k}=?"); vals.append(v)
    if "active" in d:
        sets.append("active=?"); vals.append(1 if d["active"] else 0)
    if d.get("password"):
        sets.append("password_hash=?")
        vals.append(bcrypt.hashpw(d["password"].encode(),bcrypt.gensalt()).decode())
    if sets:
        try:
            with get_db() as db:
                db.execute(f"UPDATE users SET {','.join(sets)} WHERE id=?", vals + [uid])
        except sqlite3.IntegrityError:
            return jsonify({"error": "Email уже занят другим пользователем"}), 400
        except Exception as e:
            return jsonify({"error": str(e)}), 500
    return jsonify({"ok":True})

@app.route("/api/users/<int:uid>",methods=["DELETE"])
@roles_required("superadmin")
def delete_user(uid):
    if uid==request.current_user["id"]: return jsonify({"error":"Нельзя удалить себя"}),400
    with get_db() as db: db.execute("UPDATE users SET active=0 WHERE id=?",(uid,))
    return jsonify({"ok":True})

# ─── ITEMS API ────────────────────────────────────────────────────────────────
@app.route("/api/users/simple-list")
@login_required
def get_users_simple():
    """Returns a simple list of active users for dropdowns."""
    with get_db() as db:
        rows = db.execute("SELECT id, name, department FROM users WHERE active=1 ORDER BY name").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/items/<inv_num>/reassign", methods=["POST"])
@roles_required("superadmin", "aho", "hr")
def reassign_item(inv_num):
    """Quickly reassign an asset to another employee."""
    d = request.json
    uid = d.get("user_id")
    if not uid: return jsonify({"error": "Выберите сотрудника"}), 400
    
    with get_db() as db:
        # Get target user
        user = db.execute("SELECT id, name FROM users WHERE id=?", (uid,)).fetchone()
        if not user: return jsonify({"error": "Сотрудник не найден"}), 404
        
        # Get current item info
        item = db.execute("SELECT id, employee FROM items WHERE inv_num=?", (inv_num,)).fetchone()
        if not item: return jsonify({"error": "Предмет не найден"}), 404
        
        old_val = item["employee"]
        new_val = user["name"]
        
        # Update item
        db.execute("UPDATE items SET employee=?, employee_id=?, status='Занято' WHERE inv_num=?", 
                   (new_val, uid, inv_num))
        
        # Log history
        db.execute("INSERT INTO history (item_id, user_id, user_name, action, field, old_val, new_val) VALUES (?,?,?,?,?,?,?)",
                   (item["id"], request.current_user["id"], request.current_user["name"], 
                    "Переназначение", "employee", old_val, new_val))
        
    return jsonify({"ok": True, "new_owner": new_val})

@app.route("/api/items/<inv_num>/maintenance", methods=["POST"])
@login_required
def report_maintenance(inv_num):
    """Report an issue with an asset."""
    d = request.json
    desc = d.get("description")
    if not desc: return jsonify({"error": "Опишите проблему"}), 400
    
    with get_db() as db:
        item = db.execute("SELECT id FROM items WHERE inv_num=?", (inv_num,)).fetchone()
        if not item: return jsonify({"error": "Предмет не найден"}), 404
        
        db.execute("""INSERT INTO maintenance (item_id, reported_by_id, reported_by_name, description)
                      VALUES (?,?,?,?)""",
                   (item["id"], request.current_user["id"], request.current_user["name"], desc))
        
        # Update item condition
        db.execute("UPDATE items SET condition='Требует ремонта' WHERE inv_num=?", (inv_num,))
        
    return jsonify({"ok": True})

@app.route("/api/items")
@login_required
def get_items():
    u=request.current_user
    q="SELECT * FROM items WHERE 1=1"; params=[]
    if u["role"]=="employee":
        q+=" AND (employee_id=? OR employee=?)"; params+=[u["id"],u["name"]]
    else:
        for k,col in [("room","room"),("status","status"),("category","category"),("employee","employee")]:
            v=request.args.get(k,"")
            if v: q+=f" AND {col}=?"; params.append(v)
    q+=" ORDER BY place,category"
    with get_db() as db:
        rows=db.execute(q,params).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/items",methods=["POST"])
@roles_required("superadmin","aho")
def add_item():
    if request.is_json:
        d = request.json
    else:
        d = request.form
    
    u = request.current_user
    cat = d.get("category", "Другое")
    inv = next_inv(cat)
    
    emp = d.get("employee", "—")
    default_status = "Занято" if emp != "—" else "Свободно"
    status = d.get("status", default_status)
    
    photo_name = None
    if "photo" in request.files:
        f = request.files["photo"]
        if f.filename:
            ext = os.path.splitext(f.filename)[1].lower()
            if ext in ALLOWED_EXTENSIONS:
                photo_name = f"{uuid.uuid4().hex[:12]}{ext}"
                f.save(os.path.join(UPLOADS, photo_name))

    try:
        with get_db() as db:
            cur = db.execute(
                "INSERT INTO items (place,inv_num,category,model,serial_num,room,employee,employee_id,status,condition,check_date,notes,photo,purchase_price,purchase_date,supplier,warranty_until) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (d.get("place",""), inv, cat, d.get("model",""), d.get("serial_num","—"), d.get("room",""),
                 emp, d.get("employee_id"), status,
                 d.get("condition","Хорошее"), date.today().isoformat(), d.get("notes",""), photo_name,
                 d.get("purchase_price"), d.get("purchase_date"), d.get("supplier"), d.get("warranty_until"))
            )
            item_id = cur.lastrowid
            log_h(db, item_id, "Добавлен", uid=u["id"], uname=u["name"])
    except Exception as e:
        return jsonify({"error": str(e)}),500
    
    return jsonify({"ok":True, "id": item_id, "inv_num":inv})

@app.route("/api/items/bulk",methods=["POST"])
@roles_required("superadmin","aho")
def add_bulk():
    d=request.json; u=request.current_user; inv_nums=[]
    emp = d.get("employee", "—")
    default_status = "Занято" if emp != "—" else "Свободно"
    status = d.get("status", default_status)
    try:
        with get_db() as db:
            for item in d.get("items",[]):
                inv=next_inv(item.get("category","Другое"))
                cur=db.execute("INSERT INTO items (place,inv_num,category,model,serial_num,room,employee,employee_id,status,condition,check_date,notes) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                    (d.get("place",""),inv,item.get("category","Другое"),item.get("model",""),item.get("serial_num","—"),
                     d.get("room",""),emp,d.get("employee_id"),status,
                     item.get("condition","Хорошее"),date.today().isoformat(),item.get("notes","")))
                log_h(db,cur.lastrowid,"Добавлен (пакетно)",uid=u["id"],uname=u["name"])
                inv_nums.append(inv)
    except Exception as e:
        return jsonify({"error": str(e)}),500
    return jsonify({"ok":True,"inv_nums":inv_nums,"count":len(inv_nums)})

@app.route("/api/items/<int:iid>",methods=["PUT"])
@login_required
def update_item(iid):
    u=request.current_user; d=request.json
    if u["role"]=="employee":
        with get_db() as db:
            item=db.execute("SELECT * FROM items WHERE id=?",(iid,)).fetchone()
        if not item or (str(item["employee_id"])!=str(u["id"]) and item["employee"]!=u["name"]):
            return jsonify({"error":"Нет доступа"}),403
        d={k:v for k,v in d.items() if k in ("condition","notes")}
    elif u["role"] not in ("superadmin","aho"):
        return jsonify({"error":"Нет доступа"}),403
    fields=["place","category","model","serial_num","room","employee","employee_id","status","condition","notes"]
    with get_db() as db:
        old=dict(db.execute("SELECT * FROM items WHERE id=?",(iid,)).fetchone() or {})
        sets=[f"{f}=?" for f in fields if f in d]
        vals=[d[f] for f in fields if f in d]
        if sets: db.execute(f"UPDATE items SET {','.join(sets)} WHERE id=?",vals+[iid])
        for f in fields:
            if f in d and str(d[f])!=str(old.get(f,"")):
                log_h(db,iid,"Изменено",f,str(old.get(f,"")),str(d[f]),u["id"],u["name"])
    return jsonify({"ok":True})

@app.route("/api/items/<int:iid>",methods=["DELETE"])
@login_required
def delete_item(iid):
    u = request.current_user
    if not ROLES[u["role"]].get("can_delete"):
        return jsonify({"error": "У вас нет прав на удаление"}), 403
    with get_db() as db:
        row=db.execute("SELECT photo FROM items WHERE id=?",(iid,)).fetchone()
        if row and row["photo"]:
            try: os.remove(os.path.join(UPLOADS,row["photo"]))
            except: pass
        db.execute("DELETE FROM items WHERE id=?",(iid,))
        db.execute("DELETE FROM history WHERE item_id=?",(iid,))
    return jsonify({"ok":True})

@app.route("/api/items/<int:iid>/photo",methods=["POST","DELETE"])
@login_required
def photo(iid):
    u=request.current_user
    if request.method=="DELETE":
        if u["role"] not in ("superadmin","aho"): return jsonify({"error":"Нет доступа"}),403
        with get_db() as db:
            old=db.execute("SELECT photo FROM items WHERE id=?",(iid,)).fetchone()
            if old and old["photo"]:
                try: os.remove(os.path.join(UPLOADS,old["photo"]))
                except: pass
            db.execute("UPDATE items SET photo=NULL WHERE id=?",(iid,))
            log_h(db,iid,"Фото удалено",uid=u["id"],uname=u["name"])
        return jsonify({"ok":True})
    if "photo" not in request.files: return jsonify({"ok":False})
    f=request.files["photo"]
    ext=os.path.splitext(f.filename)[1].lower() if f.filename else '.jpg'
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({"error": "Недопустимый формат файла. Разрешены: JPG, PNG, GIF, WEBP"}), 400
    name=f"{iid}_{uuid.uuid4().hex[:8]}{ext}"
    with get_db() as db:
        old=db.execute("SELECT photo,employee_id,employee FROM items WHERE id=?",(iid,)).fetchone()
        if u["role"]=="employee":
            if not old or (str(old["employee_id"])!=str(u["id"]) and old["employee"]!=u["name"]):
                return jsonify({"error":"Нет доступа"}),403
        if old and old["photo"]:
            try: os.remove(os.path.join(UPLOADS,old["photo"]))
            except: pass
    f.save(os.path.join(UPLOADS,name))
    with get_db() as db:
        db.execute("UPDATE items SET photo=? WHERE id=?",(name,iid))
        log_h(db,iid,"Фото обновлено",uid=u["id"],uname=u["name"])
    return jsonify({"ok":True,"photo":name})

@app.route("/api/items/<int:iid>/verify", methods=["POST"])
@login_required
def verify_item(iid):
    u = request.current_user
    if "photo" not in request.files:
        return jsonify({"error": "Нужно фото подтверждение"}), 400
    f = request.files["photo"]
    ext = os.path.splitext(f.filename)[1].lower() if f.filename else '.jpg'
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({"error": "Недопустимый формат файла"}), 400
    name = f"verify_{iid}_{uuid.uuid4().hex[:8]}{ext}"
    
    with get_db() as db:
        item = db.execute("SELECT * FROM items WHERE id=?", (iid,)).fetchone()
        if not item: return jsonify({"error": "Ни найдено"}), 404
        
        # Check permissions
        is_owner = str(item["employee_id"]) == str(u["id"]) or item["employee"] == u["name"]
        can_edit = ROLES[u["role"]]["can_edit"]
        if not is_owner and not can_edit:
            return jsonify({"error": "Нет доступа к верификации этого оборудования"}), 403
            
        # Save photo
        f.save(os.path.join(UPLOADS, name))
        
        # Update item
        db.execute("UPDATE items SET photo=?, check_date=? WHERE id=?", (name, date.today().isoformat(), iid))
        log_h(db, iid, "Верификация (фото-подтверждение)", uid=u["id"], uname=u["name"])
        
    return jsonify({"ok": True, "photo": name})

@app.route("/api/items/<int:iid>/history")
@login_required
def item_history(iid):
    """Returns history for a specific item."""
    with get_db() as db:
        rows = db.execute("""
            SELECT * FROM history WHERE item_id=? ORDER BY ts DESC LIMIT 50
        """, (iid,)).fetchall()
    return jsonify([dict(r) for r in rows])

# ─── HR: ISSUANCE ─────────────────────────────────────────────────────────────
@app.route("/api/issuances",methods=["GET"])
@roles_required("superadmin","aho","hr")
def list_issuances():
    with get_db() as db:
        rows=db.execute("SELECT * FROM issuances ORDER BY created_at DESC LIMIT 100").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/issuances",methods=["POST"])
@roles_required("superadmin","aho","hr")
def create_issuance():
    d=request.json; u=request.current_user
    emp_id=d.get("employee_id"); emp_name=d.get("employee_name",""); item_ids=d.get("item_ids",[])
    if not emp_id or not item_ids: return jsonify({"error":"Нужен сотрудник и список техники"}),400
    with get_db() as db:
        for iid in item_ids:
            db.execute("UPDATE items SET employee_id=?,employee=?,status='Занято' WHERE id=?",(emp_id,emp_name,iid))
            log_h(db,iid,f"Выдано: {emp_name}",uid=u["id"],uname=u["name"])
        cur=db.execute("INSERT INTO issuances (employee_id,employee_name,issued_by,issued_by_name,items_json) VALUES (?,?,?,?,?)",
            (emp_id,emp_name,u["id"],u["name"],json.dumps(item_ids)))
    return jsonify({"ok":True,"issuance_id":cur.lastrowid})

@app.route("/api/issuances/<int:iid>/confirm",methods=["POST"])
@login_required
def confirm_issuance(iid):
    u=request.current_user
    with get_db() as db:
        iss=db.execute("SELECT * FROM issuances WHERE id=?",(iid,)).fetchone()
        if not iss: return jsonify({"error":"Не найдено"}),404
        if u["role"]=="employee" and iss["employee_id"]!=u["id"]:
            return jsonify({"error":"Нет доступа"}),403
        db.execute("UPDATE issuances SET status='confirmed',confirmed_at=CURRENT_TIMESTAMP WHERE id=?",(iid,))
    return jsonify({"ok":True})

# ─── HR: RETURNS ──────────────────────────────────────────────────────────────
@app.route("/api/returns",methods=["GET"])
@roles_required("superadmin","aho","hr")
def list_returns():
    with get_db() as db:
        rows=db.execute("SELECT * FROM returns ORDER BY created_at DESC LIMIT 100").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/returns",methods=["POST"])
@roles_required("superadmin","aho","hr")
def initiate_return():
    d=request.json; u=request.current_user
    emp_id=d.get("employee_id"); emp_name=d.get("employee_name",""); item_ids=d.get("item_ids",[])
    if not emp_id or not item_ids: return jsonify({"error":"Нужен сотрудник и список техники"}),400
    with get_db() as db:
        cur=db.execute("INSERT INTO returns (employee_id,employee_name,initiated_by,initiated_by_name,items_json) VALUES (?,?,?,?,?)",
            (emp_id,emp_name,u["id"],u["name"],json.dumps(item_ids)))
    return jsonify({"ok":True,"return_id":cur.lastrowid})

@app.route("/api/returns/<int:rid>/submit-photos",methods=["POST"])
@login_required
def submit_photos(rid):
    u=request.current_user
    with get_db() as db:
        ret=db.execute("SELECT * FROM returns WHERE id=?",(rid,)).fetchone()
        if not ret: return jsonify({"error":"Не найдено"}),404
        if u["role"]=="employee" and ret["employee_id"]!=u["id"]:
            return jsonify({"error":"Нет доступа"}),403
    photos={}
    for key in request.files:
        if key.startswith("photo_"):
            item_id=key.replace("photo_",""); f=request.files[key]
            ext=os.path.splitext(f.filename)[1].lower() or ".jpg"
            name=f"ret_{rid}_{item_id}_{uuid.uuid4().hex[:6]}{ext}"
            f.save(os.path.join(UPLOADS,name)); photos[item_id]=name
    try: conditions=json.loads(request.form.get("conditions","{}") or "{}")
    except Exception as e:
        app.logger.error(f"Condition parsing error: {e}")
        conditions={}
    with get_db() as db:
        db.execute("UPDATE returns SET status='photos_submitted',photos_json=? WHERE id=?",
                   (json.dumps({"photos":photos,"conditions":conditions}),rid))
        for iid,cond in conditions.items():
            db.execute("UPDATE items SET condition=? WHERE id=?",(cond,iid))
            log_h(db,int(iid),f"Состояние при сдаче: {cond}",uid=u["id"],uname=u["name"])
    return jsonify({"ok":True})

@app.route("/api/returns/<int:rid>/accept",methods=["POST"])
@roles_required("superadmin","aho")
def accept_return(rid):
    d=request.json or {}; u=request.current_user
    with get_db() as db:
        ret=db.execute("SELECT * FROM returns WHERE id=?",(rid,)).fetchone()
        if not ret: return jsonify({"error":"Не найдено"}),404
        for iid in json.loads(ret["items_json"]):
            if d.get(str(iid)): db.execute("UPDATE items SET condition=? WHERE id=?",(d[str(iid)],iid))
            db.execute("UPDATE items SET employee_id=NULL,employee='—',status='Свободно' WHERE id=?",(iid,))
            log_h(db,iid,f"Возвращено от {ret['employee_name']}",uid=u["id"],uname=u["name"])
        db.execute("UPDATE returns SET status='completed',accepted_by=?,accepted_by_name=?,completed_at=CURRENT_TIMESTAMP WHERE id=?",
                   (u["id"],u["name"],rid))
    return jsonify({"ok":True})



# ─── STATS & EXPORT ───────────────────────────────────────────────────────────
@app.route("/api/stats")
@login_required
def stats():
    u=request.current_user
    with get_db() as db:
        if u["role"]=="employee":
            total = db.execute("SELECT COUNT(*) FROM items WHERE employee_id=? OR employee=?",(u["id"],u["name"])).fetchone()[0]
            by_cond = [dict(r) for r in db.execute("SELECT condition, COUNT(*) as cnt FROM items WHERE employee_id=? OR employee=? GROUP BY condition",(u["id"],u["name"])).fetchall()]
            return jsonify({"total":total,"occupied":total,"free":0,"broken":0,"by_cat":[],"rooms":[],"employees":[], "by_condition": by_cond})
        total   =db.execute("SELECT COUNT(*) FROM items").fetchone()[0]
        occupied=db.execute("SELECT COUNT(*) FROM items WHERE status='Занято'").fetchone()[0]
        free    =db.execute("SELECT COUNT(*) FROM items WHERE status='Свободно'").fetchone()[0]
        broken  =db.execute("SELECT COUNT(*) FROM items WHERE condition='Требует ремонта'").fetchone()[0]
        by_cat  =[dict(r) for r in db.execute("SELECT category,COUNT(*) as cnt FROM items GROUP BY category").fetchall()]
        rooms   =[r["room"] for r in db.execute("SELECT DISTINCT room FROM items WHERE room!='' ORDER BY room").fetchall()]
        emps    =[{"name":e["employee"],"count":e["cnt"]} for e in db.execute(
            "SELECT employee,COUNT(*) as cnt FROM items WHERE employee IS NOT NULL AND employee!='' AND employee!='—' GROUP BY employee ORDER BY employee"
        ).fetchall()]
        by_condition = [dict(r) for r in db.execute("SELECT condition, COUNT(*) as cnt FROM items GROUP BY condition").fetchall()]
        pending_aho = db.execute("SELECT COUNT(*) FROM dismissals WHERE status IN ('pending', 'pending_aho', 'photos_submitted')").fetchone()[0]
        pending_hr = db.execute("SELECT COUNT(*) FROM dismissals WHERE status='pending_hr'").fetchone()[0]
        
        # Pending docs for current role
        pending_docs = db.execute("SELECT COUNT(*) FROM documents WHERE status='pending' AND current_role=?", (u["role"],)).fetchone()[0]
    return jsonify({
        "total":total,"occupied":occupied,"free":free,"broken":broken,"by_cat":by_cat,"rooms":rooms,"employees":emps,
        "by_condition": by_condition,
        "pending_aho": pending_aho, "pending_hr": pending_hr,
        "pending_docs": pending_docs
    })

@app.route("/api/export")
@roles_required("superadmin","aho","auditor")
def export_excel():
    with get_db() as db:
        rows=db.execute("SELECT * FROM items ORDER BY place,category").fetchall()
    wb=Workbook(); ws=wb.active; ws.title="Инвентаризация"
    thin=Side(style="thin",color="CCCCCC"); brd=Border(left=thin,right=thin,top=thin,bottom=thin)
    headers=["№","Рабочее место","Инв. номер","Наименование","Модель","Серийный номер","Кабинет","Сотрудник","Статус","Состояние","Дата проверки","Примечания"]
    widths=[5,14,12,14,20,16,12,16,12,18,14,20]
    hfill=PatternFill("solid",start_color="1F4E79")
    ofill=PatternFill("solid",start_color="E2EFDA")
    ffill=PatternFill("solid",start_color="F2F2F2")
    rfill=PatternFill("solid",start_color="FFF2CC")
    for i,(h,w) in enumerate(zip(headers,widths),1):
        c=ws.cell(row=1,column=i,value=h)
        c.font=Font(name="Arial",bold=True,color="FFFFFF",size=10); c.fill=hfill
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        c.border=brd; ws.column_dimensions[get_column_letter(i)].width=w
    ws.row_dimensions[1].height=28; ws.freeze_panes="A2"
    for n,row in enumerate(rows,1):
        vals=[n,row["place"],row["inv_num"],row["category"],row["model"] or "",
              row["serial_num"] or "—",row["room"],row["employee"] or "—",
              row["status"],row["condition"],row["check_date"] or "",row["notes"] or ""]
        fill=ofill if row["status"]=="Занято" else(rfill if row["condition"]=="Требует ремонта" else ffill)
        for i,v in enumerate(vals,1):
            c=ws.cell(row=n+1,column=i,value=v)
            c.font=Font(name="Arial",size=9); c.fill=fill; c.border=brd
            c.alignment=Alignment(vertical="center")
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,as_attachment=True,download_name=f"Tracko_{date.today()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─── DISMISSAL WORKFLOW ───────────────────────────────────────────────────────
@app.route("/admin/dismissals")
@roles_required("superadmin","aho","hr")
def admin_dismissals_page():
    with get_db() as db:
        rows = db.execute("SELECT * FROM dismissals ORDER BY created_at DESC").fetchall()
    return render_template("admin_dismissals.html", dismissals=[dict(r) for r in rows],
                           user=request.current_user)

@app.route("/api/dismissals", methods=["POST"])
@roles_required("superadmin","aho","hr")
def create_dismissal():
    d = request.json or {}
    u = request.current_user
    emp_id = d.get("employee_id")
    if not emp_id:
        return jsonify({"error": "Не указан сотрудник"}), 400
    with get_db() as db:
        # Resolve employee by ID or Name
        emp = db.execute("SELECT * FROM users WHERE id=? OR name=?", (emp_id, emp_id)).fetchone()
        if not emp:
            return jsonify({"error": "Сотрудник не найден"}), 404
        emp_id = emp["id"]
        
        # Get all items assigned to employee
        items = db.execute(
            "SELECT id, inv_num, category, model, room FROM items WHERE employee_id=? OR employee=?",
            (emp_id, emp["name"])
        ).fetchall()
        items_json = json.dumps([dict(i) for i in items])
        # Check if active dismissal already exists
        existing = db.execute(
            "SELECT id FROM dismissals WHERE employee_id=? AND status NOT IN ('completed','cancelled')",
            (emp_id,)
        ).fetchone()
        if existing:
            return jsonify({"error": "Процесс увольнения уже запущен", "id": existing["id"]}), 400
        cur = db.execute(
            "INSERT INTO dismissals (employee_id,employee_name,employee_email,initiated_by,initiated_by_name,items_json,notes,status) VALUES (?,?,?,?,?,?,?,'pending_aho')",
            (emp_id, emp["name"], emp["email"], u["id"], u["name"], items_json, d.get("notes",""))
        )
        dis_id = cur.lastrowid
    return jsonify({"ok": True, "dismissal_id": dis_id})

@app.route("/api/dismissals/<int:did>/aho_accept", methods=["POST"])
@roles_required("superadmin","aho")
def dismissals_aho_accept(did):
    d = request.json or {}
    u = request.current_user
    with get_db() as db:
        dis = db.execute("SELECT * FROM dismissals WHERE id=?", (did,)).fetchone()
        if not dis: return jsonify({"error": "Не найдено"}), 404
        dis = dict(dis)
        if dis["status"] not in ("pending_aho", "pending", "photos_submitted"):
            return jsonify({"error": "Неверный статус"}), 400
            
        # Return all items
        items = json.loads(dis["items_json"])
        cond_map = json.loads(dis.get("item_conditions") or "{}")
        for item in items:
            iid = item["id"] if isinstance(item, dict) else item
            cond = cond_map.get(str(iid)) or d.get(str(iid)) or "Хорошее"
            if cond not in CONDITIONS: cond = "Хорошее"
            if cond == "Утеряно":
                db.execute("UPDATE items SET condition='Утеряно',employee_id=NULL,employee='—',status='Свободно' WHERE id=?",(iid,))
                log_h(db,iid,f"❌ Утеряно при увольнении: {dis['employee_name']}",uid=u["id"],uname=u["name"])
            else:
                db.execute("UPDATE items SET condition=?,employee_id=NULL,employee='—',status='Свободно' WHERE id=?",(cond,iid))
                log_h(db,iid,f"Возвращено (увольнение: {dis['employee_name']})",uid=u["id"],uname=u["name"])
            
        # Handle Signature
        sig_path = None
        if d.get("signature"):
            sig_path = _save_signature(d["signature"], f"dis_{did}_aho")

        db.execute(
            """UPDATE dismissals SET 
               status='pending_it', 
               aho_cleared=1, aho_at=CURRENT_TIMESTAMP, 
               aho_by_id=?, aho_by_name=?, aho_signature=?
               WHERE id=?""",
            (u["id"], u["name"], sig_path, did)
        )
    return jsonify({"ok": True})

@app.route("/api/dismissals/<int:did>/it_accept", methods=["POST"])
@roles_required("superadmin","aho","deputy") # IT usually handled by AHO/Sysadmin
def dismissals_it_accept(did):
    d = request.json or {}
    u = request.current_user
    with get_db() as db:
        dis = db.execute("SELECT * FROM dismissals WHERE id=?", (did,)).fetchone()
        if not dis: return jsonify({"error": "Не найдено"}), 404
        dis = dict(dis)
        if dis["status"] != "pending_it":
            return jsonify({"error": "Сначала техника должна быть принята АХО"}), 400
            
        sig_path = None
        if d.get("signature"):
            sig_path = _save_signature(d["signature"], f"dis_{did}_it")

        db.execute(
            """UPDATE dismissals SET 
               status='pending_hr', 
               it_cleared=1, it_at=CURRENT_TIMESTAMP, 
               it_by_id=?, it_by_name=?, it_signature=?
               WHERE id=?""",
            (u["id"], u["name"], sig_path, did)
        )
    return jsonify({"ok": True})

@app.route("/api/dismissals/<int:did>/hr_finalize", methods=["POST"])
@roles_required("superadmin","hr")
def dismissals_hr_finalize(did):
    u = request.current_user
    with get_db() as db:
        dis = db.execute("SELECT * FROM dismissals WHERE id=?", (did,)).fetchone()
        if not dis: return jsonify({"error": "Не найдено"}), 404
        dis = dict(dis)
        if dis["status"] != "pending_hr":
            return jsonify({"error": "Техника еще не принята АХО"}), 400
            
        # Handle Signature
        d = request.json or {}
        sig_path = None
        if d.get("signature"):
            sig_path = _save_signature(d["signature"], f"dis_{did}_hr")

        # Deactivate user
        db.execute("UPDATE users SET active=0 WHERE id=?", (dis["employee_id"],))
        db.execute(
            """UPDATE dismissals SET 
               status='completed', completed_at=CURRENT_TIMESTAMP,
               hr_at=CURRENT_TIMESTAMP, hr_by_id=?, hr_by_name=?, hr_signature=?
               WHERE id=?""",
            (u["id"], u["name"], sig_path, did)
        )
    return jsonify({"ok": True})

@app.route("/dismissal/<int:did>")
def dismissal_page(did):
    """Public page for employee to submit photos (no auth, uses dismissal token)"""
    with get_db() as db:
        dis = db.execute("SELECT * FROM dismissals WHERE id=?", (did,)).fetchone()
    if not dis:
        abort(404)
    dis = dict(dis)
    dis["items"] = json.loads(dis["items_json"] or "[]")
    dis["photos"] = json.loads(dis["photos_json"] or "{}")
    return render_template("dismissal.html", dis=dis)

@app.route("/api/dismissals/<int:did>/submit", methods=["POST"])
def submit_dismissal_photos(did):
    """Employee submits photos for each item"""
    with get_db() as db:
        dis = db.execute("SELECT * FROM dismissals WHERE id=?", (did,)).fetchone()
    if not dis:
        return jsonify({"error": "Не найдено"}), 404
    dis = dict(dis)
    if dis["status"] not in ("pending", "photos_requested"):
        return jsonify({"error": "Форма уже отправлена"}), 400
    photos = {}
    for key in request.files:
        if key.startswith("photo_"):
            item_id = key.replace("photo_", "")
            f = request.files[key]
            ext = os.path.splitext(f.filename)[1].lower() or ".jpg"
            name = f"dis_{did}_{item_id}_{uuid.uuid4().hex[:6]}{ext}"
            f.save(os.path.join(UPLOADS, name))
            photos[item_id] = name
    try:
        conditions = json.loads(request.form.get("conditions", "{}") or "{}")
    except:
        conditions = {}
    
    # Save Employee Signature
    emp_sig_path = None
    if request.form.get("signature"):
        emp_sig_path = _save_signature(request.form["signature"], f"dis_{did}_emp")

    photos_data = {"photos": photos, "conditions": conditions,
                   "comment": request.form.get("comment", ""),
                   "submitted_at": datetime.now().isoformat()}
    with get_db() as db:
        db.execute("UPDATE dismissals SET status='photos_submitted', photos_json=?, employee_signature=? WHERE id=?",
                   (json.dumps(photos_data), emp_sig_path, did))
    return jsonify({"ok": True})

@app.route("/api/dismissals/<int:did>")
@login_required
def get_dismissal(did):
    with get_db() as db:
        dis = db.execute("SELECT * FROM dismissals WHERE id=?", (did,)).fetchone()
    if not dis:
        return jsonify({"error": "Не найдено"}), 404
    dis = dict(dis)
    dis["items"] = json.loads(dis["items_json"] or "[]")
    dis["photos_data"] = json.loads(dis["photos_json"] or "{}")
    return jsonify(dis)

# Redundant confirm_dismissal removed in favor of multi-stage aho_accept/hr_finalize

@app.route("/api/dismissals/<int:did>/cancel", methods=["POST"])
@roles_required("superadmin","aho","hr")
def cancel_dismissal(did):
    u = request.current_user
    with get_db() as db:
        db.execute("UPDATE dismissals SET status='cancelled' WHERE id=?", (did,))
    return jsonify({"ok": True})

@app.route("/api/dismissals/<int:did>/request-photos", methods=["POST"])
@roles_required("superadmin","aho","hr")
def request_dismissal_photos(did):
    with get_db() as db:
        db.execute("UPDATE dismissals SET status='photos_requested' WHERE id=?", (did,))
    return jsonify({"ok": True})

# ─── SECURITY LOG ─────────────────────────────────────────────────────────────
@app.route("/api/security/login-log")
@roles_required("superadmin")
def get_login_log():
    with get_db() as db:
        rows = db.execute("""
            SELECT l.*, u.name as user_name FROM login_log l
            LEFT JOIN users u ON l.user_id = u.id
            ORDER BY l.ts DESC LIMIT 200
        """).fetchall()
    return jsonify([dict(r) for r in rows])

# ─── HISTORY (global) ────────────────────────────────────────────────────────
@app.route("/history")
@login_required
def history_page():
    u = request.current_user
    with get_db() as db:
        rows = db.execute("""
            SELECT h.*, i.inv_num, i.category, i.model
            FROM history h
            LEFT JOIN items i ON h.item_id = i.id
            ORDER BY h.ts DESC LIMIT 500
        """).fetchall()
    return render_template("history.html", rows=[dict(r) for r in rows], user=u)

@app.route("/api/history")
@login_required
def api_history():
    limit  = min(int(request.args.get("limit", 200)), 500)  # max 500
    offset = max(int(request.args.get("offset", 0)), 0)
    action_f = request.args.get("action", "")[:100]  # limit search string
    q = "SELECT h.*, i.inv_num, i.category FROM history h LEFT JOIN items i ON h.item_id=i.id WHERE 1=1"
    params = []
    if action_f:
        q += " AND h.action LIKE ?"; params.append(f"%{action_f}%")
    q += " ORDER BY h.ts DESC LIMIT ? OFFSET ?"
    params += [limit, offset]
    with get_db() as db:
        rows = db.execute(q, params).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/history/export")
@roles_required("superadmin","aho","auditor")
def export_history():
    with get_db() as db:
        rows = db.execute("SELECT h.*, i.inv_num, i.category, i.model FROM history h LEFT JOIN items i ON h.item_id=i.id ORDER BY h.ts DESC").fetchall()
    wb=Workbook(); ws=wb.active; ws.title="История изменений"
    headers=["Дата/Время","Сотрудник","Инв. №","Категория","Действие","Поле","Старое значение","Новое значение"]
    ws.append(headers)
    for row in rows:
        ws.append([row["ts"], row["user_name"], row["inv_num"] or "—", row["category"] or "—", row["action"], row["field"] or "—", str(row["old_val"] or ""), str(row["new_val"] or "")])
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,as_attachment=True,download_name=f"History_{date.today()}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─── PROFILE ──────────────────────────────────────────────────────────────────
@app.route("/profile")
@login_required
def profile_page():
    u = request.current_user
    with get_db() as db:
        my_items = db.execute("SELECT COUNT(*) FROM items WHERE employee_id=? OR employee=?", (u["id"], u["name"])).fetchone()[0]
        try:
            my_actions = db.execute("SELECT COUNT(*) FROM history WHERE user_id=?", (u["id"],)).fetchone()[0]
        except sqlite3.Error:
            my_actions = 0
    return render_template("profile.html", user=u, role_info=ROLES.get(u["role"], {}),
                           my_items=my_items, my_actions=my_actions)

@app.route("/api/profile", methods=["PUT"])
@login_required
def update_profile():
    u = request.current_user
    d = request.json or {}
    sets = []; vals = []
    if d.get("name"): sets.append("name=?"); vals.append(d["name"].strip())
    if d.get("password"):
        pw_new = d["password"]
        # Verify old password (skip if force_password_change)
        if not u.get("force_password_change"):
            old_pw = d.get("old_password", "")
            if not old_pw:
                return jsonify({"error": "Введите текущий пароль"}), 400
            if not bcrypt.checkpw(old_pw.encode(), u["password_hash"].encode()):
                return jsonify({"error": "Неверный текущий пароль"}), 403
        if len(pw_new) < 8:
            return jsonify({"error": "Минимум 8 символов"}), 400
        if not any(c.isdigit() for c in pw_new):
            return jsonify({"error": "Пароль должен содержать цифры"}), 400
        if not any(c.isalpha() for c in pw_new):
            return jsonify({"error": "Пароль должен содержать буквы"}), 400
        sets.append("password_hash=?")
        vals.append(bcrypt.hashpw(pw_new.encode(), bcrypt.gensalt()).decode())
        sets.append("token_version=COALESCE(token_version,0)+1")
        sets.append("force_password_change=0")
    if not sets:
        return jsonify({"error": "Нет данных для обновления"}), 400
    with get_db() as db:
        db.execute(f"UPDATE users SET {','.join(sets)} WHERE id=?", vals + [u["id"]])
    return jsonify({"ok": True})

# ─── NOTIFICATIONS & TELEGRAM HELPER ──────────────────────────────────────────
def send_tg_notification(user_id, message):
    """Helper to send Telegram notifications if chat_id exists."""
    with get_db() as db:
        user = db.execute("SELECT telegram_chat_id FROM users WHERE id=?", (user_id,)).fetchone()
    if user and user["telegram_chat_id"]:
        token = os.environ.get('TELEGRAM_BOT_TOKEN')
        if not token: return
        try:
            url = f"https://api.telegram.org/bot{token}/sendMessage"
            data = {"chat_id": user["telegram_chat_id"], "text": message, "parse_mode": "HTML"}
            req = urllib.request.Request(url, data=json.dumps(data).encode(), headers={'Content-Type': 'application/json'})
            with urllib.request.urlopen(req) as resp:
                pass
        except Exception as e:
            print(f"  [!] Telegram error: {e}")

@app.route("/api/maintenance/<int:rid>/action", methods=["POST"])
@roles_required("superadmin", "aho", "deputy", "accountant")
def maintenance_action(rid):
    """Approve or reject a maintenance request."""
    d = request.json
    action = d.get("action","")
    note   = d.get("note") or d.get("reason") or d.get("comment") or ""
    # normalize: 'approve'/'approved' → 'approve'; 'reject'/'rejected' → 'reject'
    if action in ("approve","approved"): action = "approve"
    elif action in ("reject","rejected"): action = "reject"
    else: return jsonify({"error": "action must be approve or reject"}), 400
    reason = note
    u = request.current_user
    
    with get_db() as db:
        req = db.execute("SELECT * FROM maintenance WHERE id=?", (rid,)).fetchone()
        if not req: return jsonify({"error": "Заявка не найдена"}), 404
        
        status = 'resolved' if action == 'approve' else 'rejected'
        db.execute("UPDATE maintenance SET status=?, resolved_by=?, resolved_at=CURRENT_TIMESTAMP, resolution=?, rejection_reason=? WHERE id=?",
                   (status, u["name"], reason if action=='approve' else '', reason if action=='reject' else '', rid))
        
        # Notify user
        msg = f"<b>🔧 Заявка на ремонт #{rid}</b>\n\nСтатус: {'✅ Одобрена' if action=='approve' else '❌ Отклонена'}\n"
        if reason: msg += f"Комментарий: {reason}"
        send_tg_notification(req["reported_by_id"], msg)
        
    return jsonify({"ok": True})

@app.route("/api/requests/<int:rid>/action", methods=["POST"])
@roles_required("superadmin", "aho", "deputy", "accountant")
def request_action(rid):
    """Approve or reject an asset purchase request."""
    d = request.json
    action = d.get("action")
    reason = d.get("reason", "")
    u = request.current_user
    
    with get_db() as db:
        req = db.execute("SELECT * FROM asset_requests WHERE id=?", (rid,)).fetchone()
        if not req: return jsonify({"error": "Заявка не найдена"}), 404
        
        status = 'approved' if action == 'approve' else 'rejected'
        db.execute("UPDATE asset_requests SET status=?, resolved_by=?, resolved_at=CURRENT_TIMESTAMP, rejection_reason=? WHERE id=?",
                   (status, u["name"], reason, rid))
        
        # Notify user
        msg = f"<b>📦 Заявка на приобретение #{rid} ({req['category']})</b>\n\nСтатус: {'✅ Одобрена' if action=='approve' else '❌ Отклонена'}\n"
        if reason: msg += f"Комментарий: {reason}"
        send_tg_notification(req["employee_id"], msg)
        
    return jsonify({"ok": True})

@app.route("/api/export/maintenance")
@roles_required("superadmin", "aho", "accountant")
def export_maintenance():
    with get_db() as db:
        rows = db.execute("""SELECT m.*, i.inv_num, i.category, i.model 
                             FROM maintenance m LEFT JOIN items i ON m.item_id=i.id 
                             ORDER BY m.created_at DESC""").fetchall()
    wb = Workbook(); ws = wb.active; ws.title = "Ремонты"
    ws.append(["ID", "Инв. №", "Категория", "Модель", "Описание проблемы", "Приоритет", "Статус", "Автор", "Кем решено", "Дата", "Результат/Причина"])
    for r in rows:
        ws.append([r["id"], r["inv_num"], r["category"], r["model"], r["description"], r["priority"], r["status"], r["reported_by_name"], r["resolved_by"], r["created_at"], r["resolution"] or r["rejection_reason"]])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"Maintenance_{date.today()}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/export/requests")
@roles_required("superadmin", "aho", "accountant", "deputy")
def export_requests():
    with get_db() as db:
        rows = db.execute("SELECT * FROM asset_requests ORDER BY created_at DESC").fetchall()
    wb = Workbook(); ws = wb.active; ws.title = "Заявки на закуп"
    ws.append(["ID", "Сотрудник", "Категория", "Причина", "Статус", "Кем решено", "Дата", "Причина отказа"])
    for r in rows:
        ws.append([r["id"], r["employee_name"], r["category"], r["reason"], r["status"], r["resolved_by"], r["created_at"], r["rejection_reason"]])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"Purchase_Requests_{date.today()}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─── EXCEL IMPORT ─────────────────────────────────────────────────────────────
@app.route("/api/items/import", methods=["POST"])
@roles_required("superadmin", "aho")
def import_excel():
    if "file" not in request.files:
        return jsonify({"error": "Нет файла"}), 400
    f = request.files["file"]
    if not f.filename.endswith((".xlsx", ".xls")):
        return jsonify({"error": "Только .xlsx файлы"}), 400
    try:
        wb = load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active
        u = request.current_user
        imported = 0; errors = []
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        # Map columns: place, room, category, model, serial_num, employee, condition, notes
        col = {}
        mapping = {
            "место": "place", "кабинет": "room", "наименование": "category",
            "категория": "category", "модель": "model", "серийный": "serial_num",
            "сотрудник": "employee", "состояние": "condition", "примечания": "notes"
        }
        for i, h in enumerate(headers):
            for key, field in mapping.items():
                if key in h: col[field] = i; break
        required = {"place", "room", "category"}
        missing = required - set(col.keys())
        if missing:
            return jsonify({"error": f"Не найдены колонки: {', '.join(missing)}. Нужны: Место, Кабинет, Наименование"}), 400
        with get_db() as db:
            for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                try:
                    def g(field, default=""):
                        idx = col.get(field)
                        val = row[idx] if idx is not None and idx < len(row) else None
                        return str(val).strip() if val is not None else default
                    place = g("place"); room = g("room"); cat = g("category", "Другое")
                    if not place or not room: continue
                    if cat not in CATEGORIES: cat = "Другое"
                    inv = next_inv(cat)
                    emp = g("employee", "—") or "—"
                    cond = g("condition", "Хорошее")
                    if cond not in CONDITIONS: cond = "Хорошее"
                    cur = db.execute(
                        "INSERT INTO items (place,inv_num,category,model,serial_num,room,employee,status,condition,check_date,notes) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                        (place, inv, cat, g("model"), g("serial_num", "—"), room, emp,
                         "Занято" if emp != "—" else "Свободно", cond, date.today().isoformat(), g("notes"))
                    )
                    log_h(db, cur.lastrowid, "Импорт из Excel", uid=u["id"], uname=u["name"])
                    imported += 1
                except Exception as e:
                    errors.append(f"Строка {ri}: {e}")
        return jsonify({"ok": True, "imported": imported, "errors": errors})
    except Exception as e:
        return jsonify({"error": f"Ошибка чтения файла: {e}"}), 400

# ─── DASHBOARD ───────────────────────────────────────────────────────────────
@app.route("/api/dashboard")
@login_required
def dashboard():
    u = request.current_user
    role = u["role"]
    res = {"role": role}
    
    with get_db() as db:
        # Common stats
        res["total_items"] = db.execute("SELECT COUNT(*) FROM items").fetchone()[0]
        res["free_count"] = db.execute("SELECT COUNT(*) FROM items WHERE status='Свободно'").fetchone()[0]
        
        # Role-specific stats
        if role in ("superadmin", "aho", "auditor"):
            res["repair_count"] = db.execute("SELECT COUNT(*) FROM items WHERE condition='Требует ремонта'").fetchone()[0]
            res["overdue_audit"] = db.execute("SELECT COUNT(*) FROM items WHERE check_date < date('now','-180 days') OR check_date IS NULL").fetchone()[0]
            res["pending_issuances"] = db.execute("SELECT COUNT(*) FROM issuances WHERE status='pending'").fetchone()[0]
            res["recent_activity"] = [dict(r) for r in db.execute("""SELECT h.action,h.user_name,h.ts,i.inv_num,i.category
                FROM history h LEFT JOIN items i ON h.item_id=i.id ORDER BY h.ts DESC LIMIT 10""").fetchall()]
        
        if role in ("superadmin", "hr", "director"):
            res["pending_dismissals"] = db.execute("SELECT COUNT(*) FROM dismissals WHERE status NOT IN ('completed','cancelled')").fetchone()[0]
            res["active_dismissals"] = [dict(r) for r in db.execute("""SELECT id,employee_name,status,created_at FROM dismissals
                WHERE status NOT IN ('completed','cancelled') ORDER BY created_at DESC LIMIT 5""").fetchall()]
            res["new_users_7d"] = db.execute("SELECT COUNT(*) FROM users WHERE created_at >= datetime('now','-7 days')").fetchone()[0]

        if role in ("superadmin", "accountant", "director", "deputy"):
            # Financial stats (assuming purchase_price exists from migrations)
            res["total_value"] = db.execute("SELECT SUM(purchase_price) FROM items").fetchone()[0] or 0
            res["monthly_spend"] = db.execute("SELECT SUM(purchase_price) FROM items WHERE purchase_date >= date('now','start of month')").fetchone()[0] or 0
            res["top_categories_value"] = [dict(r) for r in db.execute("""SELECT category, SUM(purchase_price) as total 
                FROM items GROUP BY category ORDER BY total DESC LIMIT 5""").fetchall()]

        if role == "employee":
            res["my_items_count"] = db.execute("SELECT COUNT(*) FROM items WHERE employee_id=? OR employee=?", (u["id"], u["name"])).fetchone()[0]
            res["my_pending_requests"] = db.execute("SELECT COUNT(*) FROM asset_requests WHERE employee_id=? AND status='pending'", (u["id"],)).fetchone()[0]

    return jsonify(res)

# ─── SEARCH ───────────────────────────────────────────────────────────────────
@app.route("/api/search")
@login_required
def search():
    q = request.args.get("q","").strip()
    if len(q) < 2: return jsonify({"items":[],"users":[]})
    u    = request.current_user
    like = f"%{q}%"
    with get_db() as db:
        if u["role"] == "employee":
            items = db.execute("""SELECT id,inv_num,category,model,room,employee,status,condition FROM items
                WHERE (employee_id=? OR employee=?)
                  AND (inv_num LIKE ? OR model LIKE ? OR category LIKE ? OR room LIKE ?)
                LIMIT 12""", (u["id"],u["name"],like,like,like,like)).fetchall()
            users = []
        else:
            items = db.execute("""SELECT id,inv_num,category,model,room,employee,status,condition FROM items
                WHERE inv_num LIKE ? OR model LIKE ? OR category LIKE ?
                   OR employee LIKE ? OR room LIKE ? OR serial_num LIKE ?
                ORDER BY status LIMIT 15""", (like,)*6).fetchall()
            users = db.execute("""SELECT id,name,email,role,department FROM users
                WHERE active=1 AND (name LIKE ? OR email LIKE ? OR department LIKE ?)
                LIMIT 5""", (like,like,like)).fetchall()
    return jsonify({"items":[dict(r) for r in items],"users":[dict(r) for r in users]})

# ─── BULK ACTIONS ─────────────────────────────────────────────────────────────
@app.route("/api/items/bulk-update", methods=["POST"])
@roles_required("superadmin","aho")
def bulk_update_items():
    d   = request.json or {}; u = request.current_user
    ids = d.get("ids",[])
    if not ids: return jsonify({"error":"Нет ID"}),400
    fields = {k:v for k,v in d.items() if k in ("status","condition","room","employee","employee_id") and v is not None}
    if not fields: return jsonify({"error":"Нет полей"}),400
    sets = ", ".join(f"{k}=?" for k in fields); vals = list(fields.values())
    with get_db() as db:
        for iid in ids:
            db.execute(f"UPDATE items SET {sets} WHERE id=?", vals+[iid])
            log_h(db,iid,"Массовое обновление",uid=u["id"],uname=u["name"])
    return jsonify({"ok":True,"updated":len(ids)})

@app.route("/api/items/bulk-delete", methods=["POST"])
@login_required
def bulk_delete_items():
    u = request.current_user
    if not ROLES[u["role"]].get("can_delete"):
        return jsonify({"error": "Нет прав на массовое удаление"}), 403
    d   = request.json or {}; ids = d.get("ids",[])
    if not ids: return jsonify({"error":"Нет ID"}),400
    with get_db() as db:
        for iid in ids:
            row = db.execute("SELECT photo FROM items WHERE id=?",(iid,)).fetchone()
            if row and row["photo"]:
                try: os.remove(os.path.join(UPLOADS,row["photo"]))
                except: pass
            db.execute("DELETE FROM items WHERE id=?",(iid,))
            db.execute("DELETE FROM history WHERE item_id=?",(iid,))
    return jsonify({"ok":True,"deleted":len(ids)})

@app.route("/api/dismissals/<int:did>/receive-item", methods=["POST"])
@roles_required("superadmin", "aho")
def receive_dismissal_item(did):
    """Marks a specific item in the dismissal list as physically received."""
    d = request.json
    iid = d.get("item_id")
    if not iid: return jsonify({"error": "ID предмета не указан"}), 400
    
    with get_db() as db:
        dis = db.execute("SELECT items_json FROM dismissals WHERE id=?", (did,)).fetchone()
        if not dis: return jsonify({"error": "Обходной лист не найден"}), 404
        
        items = json.loads(dis["items_json"])
        found = False
        for item in items:
            if item["id"] == iid:
                item["received"] = True
                item["received_at"] = datetime.now().isoformat()
                item["received_by"] = request.current_user["name"]
                found = True
                break
        
        if not found: return jsonify({"error": "Предмет не найден в списке"}), 404
        
        db.execute("UPDATE dismissals SET items_json=? WHERE id=?", (json.dumps(items), did))
    return jsonify({"ok": True})

@app.route("/api/dismissals/<int:did>/finalize", methods=["POST"])
@roles_required("superadmin", "hr")
def finalize_dismissal(did):
    """
    Finalizes the dismissal process ONLY if all items are received.
    """
    u = request.current_user
    with get_db() as db:
        dis = db.execute("SELECT * FROM dismissals WHERE id=?", (did,)).fetchone()
        if not dis: return jsonify({"error": "Обходной лист не найден"}), 404
        if dis["status"] == "completed": return jsonify({"error": "Уже завершено"}), 400
        
        # Check if all items are received
        items = json.loads(dis["items_json"])
        for item in items:
            if not item.get("received"):
                return jsonify({"error": f"Сначала примите все вещи (не принят: {item.get('inv_num')})"}), 400
        
        emp_id = dis["employee_id"]
        emp_name = dis["employee_name"]
        
        # Deactivate User and Release Assets
        db.execute("UPDATE users SET active=0 WHERE id=?", (emp_id,))
        asset_items = db.execute("SELECT id FROM items WHERE employee_id=? OR employee=?", (emp_id, emp_name)).fetchall()
        for item in asset_items:
            db.execute("UPDATE items SET employee_id=NULL, employee='—', status='Свободно' WHERE id=?", (item["id"],))
            log_h(db, item["id"], f"Освобождено (увольнение {emp_name})", uid=u["id"], uname=u["name"])
            
        db.execute("UPDATE dismissals SET status='completed', completed_at=CURRENT_TIMESTAMP, confirmed_by=?, confirmed_by_name=? WHERE id=?",
                   (u["id"], u["name"], did))
        
        send_tg_notification(dis["initiated_by"], f"<b>✅ Обходной лист закрыт</b>\nСотрудник: {emp_name}\nВсе вещи приняты АХО.")
        
    return jsonify({"ok": True})

@app.route("/api/items/audit", methods=["POST"])
@roles_required("superadmin","aho","auditor")
def audit_items():
    d = request.json or {}; u = request.current_user; ids = d.get("ids",[])
    if not ids: return jsonify({"error":"Нет ID"}),400
    with get_db() as db:
        for iid in ids:
            db.execute("UPDATE items SET check_date=? WHERE id=?",(date.today().isoformat(),iid))
            log_h(db,iid,"Аудит проведён",uid=u["id"],uname=u["name"])
    return jsonify({"ok":True,"audited":len(ids)})

# ─── ONBOARDING ───────────────────────────────────────────────────────────────
@app.route("/api/onboarding/free-items")
@roles_required("superadmin","aho","hr")
def free_items_for_onboarding():
    room = request.args.get("room","")
    q    = "SELECT * FROM items WHERE status='Свободно'"
    p    = []
    if room: q += " AND room=?"; p.append(room)
    q += " ORDER BY category,room,inv_num"
    with get_db() as db:
        items = db.execute(q,p).fetchall()
    grouped = {}
    for item in items:
        cat = item["category"]
        if cat not in grouped: grouped[cat] = []
        grouped[cat].append(dict(item))
    return jsonify(grouped)

# ─── REPORTS ─────────────────────────────────────────────────────────────────
@app.route("/api/reports/summary")
@roles_required("superadmin","aho","auditor")
def report_summary():
    with get_db() as db:
        by_room = db.execute("""SELECT room,COUNT(*) as total,
            SUM(CASE WHEN status='Занято' THEN 1 ELSE 0 END) as occupied,
            SUM(CASE WHEN status='Свободно' THEN 1 ELSE 0 END) as free,
            SUM(CASE WHEN condition='Требует ремонта' THEN 1 ELSE 0 END) as repair
            FROM items GROUP BY room ORDER BY room""").fetchall()
        by_cat = db.execute("""SELECT category,COUNT(*) as total,
            SUM(CASE WHEN status='Занято' THEN 1 ELSE 0 END) as occupied,
            SUM(CASE WHEN condition='Требует ремонта' THEN 1 ELSE 0 END) as repair
            FROM items GROUP BY category ORDER BY total DESC""").fetchall()
        by_emp = db.execute("""SELECT employee,COUNT(*) as cnt FROM items
            WHERE employee IS NOT NULL AND employee!='—' AND employee!=''
            GROUP BY employee ORDER BY cnt DESC LIMIT 20""").fetchall()
        total = db.execute("SELECT COUNT(*) FROM items").fetchone()[0]
        users_active = db.execute("SELECT COUNT(*) FROM users WHERE active=1").fetchone()[0]
    return jsonify({
        "total_items": total, "active_users": users_active,
        "by_room": [dict(r) for r in by_room],
        "by_category": [dict(r) for r in by_cat],
        "top_employees": [dict(r) for r in by_emp],
    })

@app.route("/api/reports/export-full")
@roles_required("superadmin","aho","auditor")
def export_full_report():
    with get_db() as db:
        items      = db.execute("SELECT * FROM items ORDER BY room,place,category").fetchall()
        hist       = db.execute("""SELECT h.ts,h.user_name,h.action,h.field,h.old_val,h.new_val,i.inv_num,i.category
            FROM history h LEFT JOIN items i ON h.item_id=i.id ORDER BY h.ts DESC LIMIT 2000""").fetchall()
        dismissals = db.execute("SELECT * FROM dismissals ORDER BY created_at DESC").fetchall()
        issuances  = db.execute("SELECT * FROM issuances ORDER BY created_at DESC").fetchall()
    wb   = Workbook()
    thin = Side(style="thin",color="CCCCCC"); brd = Border(left=thin,right=thin,top=thin,bottom=thin)
    hfill= PatternFill("solid",start_color="1F4E79")
    ffill= PatternFill("solid",start_color="F2F2F2")
    ofill= PatternFill("solid",start_color="E2EFDA")
    rfill= PatternFill("solid",start_color="FFF2CC")
    def ws_hdr(ws,headers,widths):
        for i,(h,w) in enumerate(zip(headers,widths),1):
            c=ws.cell(row=1,column=i,value=h)
            c.font=Font(name="Arial",bold=True,color="FFFFFF",size=10); c.fill=hfill
            c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=brd
            ws.column_dimensions[get_column_letter(i)].width=w
        ws.row_dimensions[1].height=26; ws.freeze_panes="A2"
    def ws_row(ws,n,vals,fill):
        for i,v in enumerate(vals,1):
            c=ws.cell(row=n,column=i,value=v); c.font=Font(name="Arial",size=9)
            c.fill=fill; c.border=brd; c.alignment=Alignment(vertical="center")
    ws1=wb.active; ws1.title="Активы"
    ws_hdr(ws1,["№","Инв.№","Категория","Модель","Серийный №","Кабинет","Место","Статус","Состояние","Сотрудник","Дата проверки","Примечания"],
              [4,10,12,18,14,12,10,10,14,16,12,18])
    for n,row in enumerate(items,1):
        fill=ofill if row["status"]=="Занято" else(rfill if row["condition"]=="Требует ремонта" else ffill)
        ws_row(ws1,n+1,[n,row["inv_num"],row["category"],row["model"] or "",row["serial_num"] or "—",
                        row["room"],row["place"],row["status"],row["condition"],
                        row["employee"] or "—",row["check_date"] or "",row["notes"] or ""],fill)
    ws2=wb.create_sheet("История")
    ws_hdr(ws2,["Дата/Время","Пользователь","Инв.№","Категория","Действие","Поле","Было","Стало"],[16,16,10,12,20,12,16,16])
    for n,row in enumerate(hist,1):
        ws_row(ws2,n+1,[row["ts"],row["user_name"] or "—",row["inv_num"] or "—",row["category"] or "—",
                        row["action"],row["field"] or "—",str(row["old_val"] or ""),str(row["new_val"] or "")],ffill)
    ws3=wb.create_sheet("Увольнения")
    ws_hdr(ws3,["Дата","Сотрудник","Email","Кто инициировал","Статус","Завершено","Примечания"],[16,20,22,20,14,16,20])
    for n,row in enumerate(dismissals,1):
        ws_row(ws3,n+1,[row["created_at"],row["employee_name"],row["employee_email"] or "",
                        row["initiated_by_name"],row["status"],row["completed_at"] or "—",row["notes"] or ""],ffill)
    ws4=wb.create_sheet("Выдачи")
    ws_hdr(ws4,["Дата","Сотрудник","Кто выдал","Статус","Подтверждено"],[16,20,20,12,16])
    for n,row in enumerate(issuances,1):
        ws_row(ws4,n+1,[row["created_at"],row["employee_name"],row["issued_by_name"],
                        row["status"],row["confirmed_at"] or "—"],ffill)
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,as_attachment=True,
                     download_name=f"Tracko_FullReport_{date.today()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─── ASSIGN VIA QR ────────────────────────────────────────────────────────────
@app.route("/api/users/active")
@login_required
def get_active_users():
    """Return list of active users for QR assign dropdown."""
    with get_db() as db:
        rows = db.execute(
            "SELECT id, name, email, role, department FROM users WHERE active=1 ORDER BY name"
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/items/<int:iid>/assign", methods=["POST"])
@roles_required("superadmin", "aho", "hr")
def assign_item(iid):
    """Quickly assign or release an item via QR-scan page."""
    d = request.json or {}
    u = request.current_user
    emp_id   = d.get("employee_id")   # None → release
    emp_name = d.get("employee_name", "")

    with get_db() as db:
        item = db.execute("SELECT * FROM items WHERE id=?", (iid,)).fetchone()
        if not item:
            return jsonify({"error": "Предмет не найден"}), 404

        old_emp = item["employee"] or "—"

        if emp_id:
            # Assign to employee
            emp = db.execute("SELECT * FROM users WHERE id=? AND active=1", (emp_id,)).fetchone()
            if not emp:
                return jsonify({"error": "Сотрудник не найден"}), 404
            emp_name = emp["name"]
            db.execute(
                "UPDATE items SET employee_id=?, employee=?, status='Занято' WHERE id=?",
                (emp_id, emp_name, iid)
            )
            action = f"QR-назначение: {old_emp} → {emp_name}"
        else:
            # Release item
            db.execute(
                "UPDATE items SET employee_id=NULL, employee='—', status='Свободно' WHERE id=?",
                (iid,)
            )
            action = f"QR-освобождение: {old_emp} → Свободно"

        log_h(db, iid, action, uid=u["id"], uname=u["name"])

    return jsonify({"ok": True, "action": action,
                    "employee": emp_name if emp_id else "—",
                    "status": "Занято" if emp_id else "Свободно"})

# ─── HEALTH ──────────────────────────────────────────────────────────────────
@app.route("/api/health")
@login_required
def health():
    """Health check — requires auth to prevent info disclosure"""
    try:
        with get_db() as db:
            db.execute("SELECT 1").fetchone()
        return jsonify({"status": "ok"})
    except Exception:
        return jsonify({"status": "error"}), 500


# ─── TRANSFER ASSET (новый ответственный) ────────────────────────────────────
@app.route("/api/items/<int:iid>/transfer", methods=["POST"])
@roles_required("superadmin", "aho", "hr")
def transfer_item(iid):
    """Переназначить актив другому сотруднику с записью в историю."""
    d = request.json or {}
    u = request.current_user
    new_emp_id   = d.get("employee_id")
    note         = (d.get("note") or "").strip()[:500]

    with get_db() as db:
        item = db.execute("SELECT * FROM items WHERE id=?", (iid,)).fetchone()
        if not item:
            return jsonify({"error": "Актив не найден"}), 404
        old_emp = item["employee"] or "—"

        if new_emp_id:
            emp = db.execute("SELECT * FROM users WHERE id=? AND active=1", (new_emp_id,)).fetchone()
            if not emp:
                return jsonify({"error": "Сотрудник не найден"}), 404
            new_name = emp["name"]
            db.execute(
                "UPDATE items SET employee_id=?, employee=?, status='Занято' WHERE id=?",
                (new_emp_id, new_name, iid)
            )
            action = f"Передача: {old_emp} → {new_name}"
        else:
            new_name = "—"
            db.execute(
                "UPDATE items SET employee_id=NULL, employee='—', status='Свободно' WHERE id=?",
                (iid,)
            )
            action = f"Освобождение: {old_emp} → Свободно"

        if note:
            action += f" | {note}"
        log_h(db, iid, action, "employee", old_emp, new_name, u["id"], u["name"])

    return jsonify({"ok": True, "action": action})


# ─── DEPARTMENTS API ──────────────────────────────────────────────────────────
@app.route("/api/departments")
@login_required
def get_departments():
    """Список отделов из пользователей + статистика по активам."""
    with get_db() as db:
        depts = db.execute(
            "SELECT department, COUNT(*) as user_count FROM users WHERE active=1 AND department IS NOT NULL AND department!='' GROUP BY department ORDER BY department"
        ).fetchall()
        result = []
        for d in depts:
            items = db.execute(
                """SELECT COUNT(*) as total,
                   SUM(CASE WHEN status='Занято' THEN 1 ELSE 0 END) as occupied,
                   SUM(CASE WHEN condition='Требует ремонта' THEN 1 ELSE 0 END) as repair
                   FROM items i
                   JOIN users u ON u.name = i.employee
                   WHERE u.department=? AND u.active=1""",
                (d["department"],)
            ).fetchone()
            result.append({
                "name": d["department"],
                "user_count": d["user_count"],
                "total_items": items["total"] or 0,
                "occupied": items["occupied"] or 0,
                "repair": items["repair"] or 0,
            })
    return jsonify(result)


@app.route("/api/departments/<path:dept>/items")
@roles_required("superadmin", "aho", "hr", "auditor")
def get_dept_items(dept):
    """Все активы сотрудников отдела."""
    with get_db() as db:
        rows = db.execute(
            """SELECT i.* FROM items i
               JOIN users u ON u.name = i.employee AND u.active=1
               WHERE u.department=?
               ORDER BY u.name, i.category""",
            (dept,)
        ).fetchall()
    return jsonify([dict(r) for r in rows])


# ─── WEBHOOK / INTEGRATION LAYER ─────────────────────────────────────────────
@app.route("/api/webhooks/test", methods=["POST"])
@roles_required("superadmin")
def webhook_test():
    """Проверить webhook endpoint."""
    d = request.json or {}
    url = d.get("url", "")
    if not url.startswith("https://") and not url.startswith("http://"):
        return jsonify({"error": "Некорректный URL"}), 400
    try:
        payload = json.dumps({"event": "ping", "source": "tracko"}).encode()
        req = urllib.request.Request(url, data=payload,
                                      headers={"Content-Type": "application/json"},
                                      method="POST")
        urllib.request.urlopen(req, timeout=5)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/integrations/uzinfocom/sync", methods=["POST"])
@roles_required("superadmin")
def uzinfocom_sync():
    """
    Заглушка для интеграции с Uzinfocom CRM / HR.
    Принимает список сотрудников и синхронизирует с users.
    Формат: {"employees": [{"name":"...", "email":"...", "department":"..."}]}
    """
    d = request.json or {}
    employees = d.get("employees", [])
    if not employees:
        return jsonify({"error": "Нет сотрудников"}), 400
    u = request.current_user
    created = 0; updated = 0; errors = []
    with get_db() as db:
        for emp in employees:
            name  = (emp.get("name") or "").strip()
            email = (emp.get("email") or "").strip().lower()
            dept  = (emp.get("department") or "").strip()
            if not name or not email:
                errors.append(f"Пропуск: нет имени или email — {emp}")
                continue
            try:
                existing = db.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone()
                if existing:
                    db.execute("UPDATE users SET name=?, department=?, active=1 WHERE email=?",
                               (name, dept, email))
                    updated += 1
                else:
                    tmp_pw = bcrypt.hashpw(secrets.token_hex(16).encode(), bcrypt.gensalt()).decode()
                    db.execute("INSERT INTO users (name,email,password_hash,role,department) VALUES (?,?,?,?,?)",
                               (name, email, tmp_pw, "employee", dept))
                    created += 1
            except Exception as e:
                errors.append(f"{email}: {e}")
    return jsonify({"ok": True, "created": created, "updated": updated, "errors": errors})


# ─── EXPORT DEPARTMENT ────────────────────────────────────────────────────────
@app.route("/api/departments/<path:dept>/export")
@roles_required("superadmin", "aho", "auditor")
def export_dept(dept):
    """Excel-экспорт по отделу."""
    with get_db() as db:
        rows = db.execute(
            """SELECT i.inv_num, i.category, i.model, i.serial_num,
                      i.room, i.place, i.status, i.condition, i.check_date,
                      i.notes, u.name as emp_name, u.email as emp_email
               FROM items i
               JOIN users u ON u.name = i.employee AND u.active=1
               WHERE u.department=?
               ORDER BY u.name, i.category""",
            (dept,)
        ).fetchall()
    wb = Workbook(); ws = wb.active; ws.title = dept[:31]
    thin = Side(style="thin", color="CCCCCC")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfill = PatternFill("solid", start_color="1F4E79")
    headers = ["Инв.№","Категория","Модель","Серийный №","Кабинет","Место",
               "Статус","Состояние","Дата проверки","Примечания","Сотрудник","Email"]
    widths  = [10,12,18,14,12,10,10,14,12,18,18,22]
    for i,(h,w) in enumerate(zip(headers,widths),1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(name="Arial",bold=True,color="FFFFFF",size=10); c.fill = hfill
        c.alignment = Alignment(horizontal="center",vertical="center"); c.border = brd
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 26; ws.freeze_panes = "A2"
    ffill = PatternFill("solid",start_color="F2F2F2")
    ofill = PatternFill("solid",start_color="E2EFDA")
    for n, row in enumerate(rows, 1):
        fill = ofill if row["status"]=="Занято" else ffill
        vals = [row["inv_num"],row["category"],row["model"] or "",
                row["serial_num"] or "—",row["room"],row["place"],
                row["status"],row["condition"],row["check_date"] or "",
                row["notes"] or "",row["emp_name"],row["emp_email"] or ""]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=n+1, column=i, value=v)
            c.font=Font(name="Arial",size=9); c.fill=fill; c.border=brd
            c.alignment=Alignment(vertical="center")
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = dept.replace("/","_").replace(" ","_")
    return send_file(buf, as_attachment=True,
                     download_name=f"Dept_{fname}_{date.today()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ─── ASSET REQUESTS (сотрудник просит технику) ───────────────────────────────
@app.route("/api/requests", methods=["GET"])
@login_required
def list_requests():
    u = request.current_user
    with get_db() as db:
        if u["role"] == "employee":
            rows = db.execute(
                "SELECT * FROM asset_requests WHERE employee_id=? ORDER BY created_at DESC LIMIT 50",
                (u["id"],)
            ).fetchall()
        else:
            rows = db.execute(
                "SELECT * FROM asset_requests ORDER BY created_at DESC LIMIT 100"
            ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/requests", methods=["POST"])
@login_required
def create_request():
    d = request.json or {}
    u = request.current_user
    category = d.get("category", "Другое")
    reason   = (d.get("reason") or "").strip()[:500]
    if not category:
        return jsonify({"error": "Укажите категорию"}), 400
    with get_db() as db:
        cur = db.execute(
            "INSERT INTO asset_requests (employee_id, employee_name, category, reason, status) VALUES (?,?,?,?,?)",
            (u["id"], u["name"], category, reason, "pending")
        )
    return jsonify({"ok": True, "id": cur.lastrowid})


@app.route("/api/requests/<int:rid>", methods=["PUT"])
@roles_required("superadmin", "aho", "hr")
def update_request(rid):
    d  = request.json or {}
    u  = request.current_user
    st = d.get("status")  # approved / rejected / completed
    if st not in ("approved", "rejected", "completed"):
        return jsonify({"error": "Некорректный статус"}), 400
    with get_db() as db:
        db.execute(
            "UPDATE asset_requests SET status=?, resolved_by=?, resolved_at=CURRENT_TIMESTAMP WHERE id=?",
            (st, u["name"], rid)
        )
    return jsonify({"ok": True})


# ─── MAINTENANCE LOG (заявки на ремонт) ──────────────────────────────────────
@app.route("/api/maintenance", methods=["GET"])
@login_required
def list_maintenance():
    with get_db() as db:
        rows = db.execute(
            """SELECT m.*, i.inv_num, i.category, i.model, i.room
               FROM maintenance m
               LEFT JOIN items i ON m.item_id = i.id
               ORDER BY m.created_at DESC LIMIT 100"""
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/maintenance", methods=["POST"])
@login_required
def create_maintenance():
    d = request.json or {}
    u = request.current_user
    item_id = d.get("item_id")
    if not item_id:
        return jsonify({"error": "Укажите актив"}), 400
    description = (d.get("description") or "").strip()[:1000]
    priority    = d.get("priority", "medium")
    if priority not in ("low", "medium", "high", "critical"):
        priority = "medium"
    with get_db() as db:
        item = db.execute("SELECT * FROM items WHERE id=?", (item_id,)).fetchone()
        if not item:
            return jsonify({"error": "Актив не найден"}), 404
        db.execute("UPDATE items SET condition='Требует ремонта' WHERE id=?", (item_id,))
        log_h(db, item_id, "Заявка на ремонт", uid=u["id"], uname=u["name"])
        cur = db.execute(
            "INSERT INTO maintenance (item_id, reported_by_id, reported_by_name, description, priority) VALUES (?,?,?,?,?)",
            (item_id, u["id"], u["name"], description, priority)
        )
    return jsonify({"ok": True, "id": cur.lastrowid})


@app.route("/api/maintenance/<int:mid>", methods=["PUT"])
@roles_required("superadmin", "aho")
def update_maintenance(mid):
    d  = request.json or {}
    u  = request.current_user
    st = d.get("status")
    if st not in ("in_progress","completed","cancelled"):
        return jsonify({"error": "Некорректный статус"}), 400
    resolution = (d.get("resolution") or d.get("note") or d.get("comment") or "").strip()[:500]
    with get_db() as db:
        m = db.execute("SELECT * FROM maintenance WHERE id=?", (mid,)).fetchone()
        if not m:
            return jsonify({"error": "Не найдено"}), 404
        db.execute(
            "UPDATE maintenance SET status=?, resolved_by=?, resolved_at=CURRENT_TIMESTAMP, resolution=? WHERE id=?",
            (st, u["name"], resolution, mid)
        )
        if st == "completed":
            db.execute("UPDATE items SET condition='Хорошее' WHERE id=?", (m["item_id"],))
            log_h(db, m["item_id"], "Ремонт завершён", uid=u["id"], uname=u["name"])
    return jsonify({"ok": True})


# ─── ANALYTICS ────────────────────────────────────────────────────────────────
@app.route("/api/analytics")
@roles_required("superadmin", "aho", "auditor")
def analytics():
    """Расширенная аналитика для дашборда."""
    with get_db() as db:
        # Стоимость потенциального простоя (кол-во "Требует ремонта")
        repair_count = db.execute("SELECT COUNT(*) FROM items WHERE condition='Требует ремонта'").fetchone()[0]
        
        # Utilization по отделам
        dept_util = db.execute("""
            SELECT u.department,
                   COUNT(i.id) as total,
                   SUM(CASE WHEN i.status='Занято' THEN 1 ELSE 0 END) as occupied
            FROM users u
            LEFT JOIN items i ON i.employee_id = u.id OR i.employee = u.name
            WHERE u.active=1 AND u.department IS NOT NULL AND u.department != ''
            GROUP BY u.department
            ORDER BY total DESC
        """).fetchall()
        
        # Активность за 30 дней
        activity_30d = db.execute("""
            SELECT DATE(ts) as day, COUNT(*) as cnt
            FROM history
            WHERE ts >= datetime('now','-30 days')
            GROUP BY DATE(ts)
            ORDER BY day
        """).fetchall()
        
        # Топ категории требующие внимания
        attention = db.execute("""
            SELECT category,
                   COUNT(*) as total,
                   SUM(CASE WHEN condition='Требует ремонта' THEN 1 ELSE 0 END) as repair,
                   SUM(CASE WHEN check_date < date('now','-180 days') OR check_date IS NULL THEN 1 ELSE 0 END) as overdue
            FROM items
            GROUP BY category
            HAVING repair > 0 OR overdue > 0
            ORDER BY (repair + overdue) DESC
            LIMIT 10
        """).fetchall()
        
        # ── Financials for Director ──
        total_val = db.execute("SELECT SUM(purchase_price) FROM items").fetchone()[0] or 0
        
        # ── Spending by Dept ──
        dept_spending = db.execute("""
            SELECT u.department, SUM(i.purchase_price) as spent
            FROM users u
            JOIN items i ON i.employee_id = u.id
            WHERE u.active=1 AND u.department IS NOT NULL AND i.purchase_price IS NOT NULL
            GROUP BY u.department
            ORDER BY spent DESC
        """).fetchall()

        # ── Strategic Health: Toxic Assets (items with > 3 maintenance records) ──
        toxic_assets = db.execute("""
            SELECT i.category, i.model, i.inv_num, COUNT(m.id) as repair_count
            FROM items i
            JOIN maintenance m ON m.item_id = i.id
            WHERE m.status = 'resolved'
            GROUP BY i.id
            HAVING repair_count >= 3
            ORDER BY repair_count DESC
            LIMIT 5
        """).fetchall()

        # ── Strategic Health: EOL Assets (End of Life - items > 3 years old) ──
        eol_assets = db.execute("""
            SELECT category, model, inv_num, purchase_date
            FROM items
            WHERE purchase_date < date('now', '-3 years')
            LIMIT 5
        """).fetchall()

        # ── Pending Counts ──
        pending_dismiss = db.execute("SELECT COUNT(*) FROM dismissals WHERE status NOT IN ('completed','cancelled')").fetchone()[0]
        pending_maint   = db.execute("SELECT COUNT(*) FROM maintenance WHERE status='pending'").fetchone()[0]
        pending_req     = db.execute("SELECT COUNT(*) FROM asset_requests WHERE status='pending'").fetchone()[0]
        
        # Accurate Total pending docs across all workflows
        p_issuances = db.execute("SELECT COUNT(*) FROM issuances WHERE status='pending'").fetchone()[0]
        p_returns   = db.execute("SELECT COUNT(*) FROM returns WHERE status='pending'").fetchone()[0]
        p_dismiss   = pending_dismiss
        pending_docs = p_issuances + p_returns + p_dismiss

    return jsonify({
        "repair_count": repair_count,
        "total_value": total_val,
        "dept_spending": [dict(r) for r in dept_spending],
        "toxic_assets": [dict(r) for r in toxic_assets],
        "eol_assets": [dict(r) for r in eol_assets],
        "pending_dismissals": pending_dismiss,
        "pending_maintenance": pending_maint,
        "pending_requests": pending_req,
        "pending_docs": pending_docs,
        "dept_utilization": [dict(r) for r in dept_util],
        "activity_30d": [dict(r) for r in activity_30d],
        "attention_items": [dict(r) for r in attention],
    })


# ─── AUDIT EXPORT SIGNED ─────────────────────────────────────────────────────
@app.route("/api/audit/sign", methods=["POST"])
@roles_required("superadmin", "aho", "auditor")
def sign_audit():
    """Подписать акт инвентаризации — сохраняет timestamp и подписавшего."""
    u  = request.current_user
    d  = request.json or {}
    ids = d.get("ids", [])  # item ids included in audit
    note = (d.get("note") or "").strip()[:500]
    with get_db() as db:
        today = date.today().isoformat()
        for iid in ids:
            db.execute("UPDATE items SET check_date=? WHERE id=?", (today, iid))
        cur = db.execute(
            "INSERT INTO audit_log (signed_by_id, signed_by_name, item_count, note) VALUES (?,?,?,?)",
            (u["id"], u["name"], len(ids), note)
        )
    return jsonify({"ok": True, "audit_id": cur.lastrowid, "date": today, "count": len(ids)})


@app.route("/api/audit/history")
@roles_required("superadmin", "aho", "auditor")
def audit_history():
    with get_db() as db:
        rows = db.execute(
            "SELECT * FROM audit_log ORDER BY created_at DESC LIMIT 50"
        ).fetchall()
    return jsonify([dict(r) for r in rows])



# ═══════════════════════════════════════════════════════════════════════════
#  ДОКУМЕНТООБОРОТ — DOCFLOW MODULE
#  Цепочка согласования: Сотрудник → АХО → Зам.Директора → Ген.Директор → Бухгалтер
# ═══════════════════════════════════════════════════════════════════════════

# ─── РОЛИ ДОКУМЕНТООБОРОТА ─────────────────────────────────────────────────
# employee   → создаёт заявку
# aho        → первое согласование (АХО/IT)
# deputy     → согласование зам. директора  [НОВАЯ РОЛЬ]
# director   → финальное утверждение ген. директора [НОВАЯ РОЛЬ]
# accountant → бухгалтер — ставит печать, закрывает [НОВАЯ РОЛЬ]
# superadmin → видит всё, может всё

# Маппинг ролей → уровень согласования
APPROVAL_CHAIN = {
    "doc_request": [          # Заявка на технику
        {"step": 1, "role": "aho",        "label": "АХО / IT"},
        {"step": 2, "role": "deputy",     "label": "Зам. Директора"},
        {"step": 3, "role": "director",   "label": "Ген. Директор"},
        {"step": 4, "role": "accountant", "label": "Бухгалтер"},
    ],
    "write_off": [            # Списание техники
        {"step": 1, "role": "aho",        "label": "АХО / IT"},
        {"step": 2, "role": "director",   "label": "Ген. Директор"},
        {"step": 3, "role": "accountant", "label": "Бухгалтер"},
    ],
    "repair": [               # Заявка на ремонт
        {"step": 1, "role": "aho",        "label": "АХО / IT"},
        {"step": 2, "role": "director",   "label": "Ген. Директор"},
    ],
    "transfer": [             # Передача техники
        {"step": 1, "role": "aho",        "label": "АХО / IT"},
        {"step": 2, "role": "deputy",     "label": "Зам. Директора"},
    ],
}

DOC_TYPES = {
    "doc_request": "Заявка на технику",
    "write_off":   "Списание техники",
    "repair":      "Заявка на ремонт",
    "transfer":    "Передача техники",
}

def _next_doc_step(doc_type, current_step):
    """Вернуть следующий шаг и роль согласования."""
    chain = APPROVAL_CHAIN.get(doc_type, [])
    nxt = [s for s in chain if s["step"] > current_step]
    return nxt[0] if nxt else None

def _doc_status_label(status):
    return {
        "draft":    "Черновик",
        "pending":  "На согласовании",
        "approved": "Утверждено",
        "rejected": "Отклонено",
        "printed":  "Распечатано",
        "closed":   "Закрыто",
    }.get(status, status)


# ─── INIT DOCFLOW TABLES ──────────────────────────────────────────────────
def _init_docflow_tables(db):
    """Создать таблицы документооборота (вызывается из init_db)."""
    db.execute("""
    CREATE TABLE IF NOT EXISTS documents (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        doc_number  TEXT UNIQUE,
        doc_type    TEXT NOT NULL,
        title       TEXT NOT NULL,
        description TEXT,
        priority    TEXT DEFAULT 'normal',
        status      TEXT DEFAULT 'draft',
        current_step INTEGER DEFAULT 0,
        current_role TEXT,
        created_by_id   INTEGER,
        created_by_name TEXT,
        item_id     INTEGER,
        item_inv    TEXT,
        department  TEXT,
        amount      REAL,
        attachments TEXT DEFAULT '[]',
        created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        deadline    DATE,
        closed_at   TIMESTAMP,
        employee_id INTEGER,
        employee_name TEXT,
        signature   TEXT
    )""")
    db.execute("""
    CREATE TABLE IF NOT EXISTS doc_approvals (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        doc_id      INTEGER NOT NULL,
        step        INTEGER NOT NULL,
        role        TEXT NOT NULL,
        role_label  TEXT,
        approver_id   INTEGER,
        approver_name TEXT,
        action      TEXT,
        comment     TEXT,
        acted_at    TIMESTAMP,
        created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (doc_id) REFERENCES documents(id)
    )""")
    db.execute("""
    CREATE TABLE IF NOT EXISTS doc_comments (
        id       INTEGER PRIMARY KEY AUTOINCREMENT,
        doc_id   INTEGER NOT NULL,
        user_id  INTEGER,
        user_name TEXT,
        user_role TEXT,
        text     TEXT NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (doc_id) REFERENCES documents(id)
    )""")
    # Add docflow roles to existing users table
    try:
        db.execute("ALTER TABLE users ADD COLUMN doc_role TEXT")
    except: pass
    try:
        db.execute("ALTER TABLE users ADD COLUMN department TEXT")
    except: pass

def _gen_doc_number(db, doc_type):
    """Сгенерировать номер документа: ЗАЯ-2025-0001"""
    prefixes = {"doc_request":"ЗАЯ","write_off":"СПС","repair":"РЕМ","transfer":"ПЕР"}
    pref = prefixes.get(doc_type, "ДОК")
    year = date.today().year
    cnt = db.execute(
        "SELECT COUNT(*) FROM documents WHERE doc_type=? AND strftime('%Y',created_at)=?",
        (doc_type, str(year))
    ).fetchone()[0] + 1
    return f"{pref}-{year}-{cnt:04d}"


# ─── ROUTES ──────────────────────────────────────────────────────────────

@app.route("/documents")
@login_required
def documents_page():
    u = request.current_user
    with get_db() as db:
        # Count pending for badge
        if u["role"] in ("superadmin","aho","deputy","director","accountant"):
            pending = db.execute(
                "SELECT COUNT(*) FROM documents WHERE status='pending' AND current_role=?",
                (u["role"],)
            ).fetchone()[0]
        else:
            pending = db.execute(
                "SELECT COUNT(*) FROM documents WHERE created_by_id=? AND status NOT IN ('closed','rejected')",
                (u["id"],)
            ).fetchone()[0]
    return render_template("documents.html",
        user=u, current_user=u,
        role_info=ROLES.get(u["role"],{}),
        doc_types=DOC_TYPES,
        approval_chain=APPROVAL_CHAIN,
        pending_count=pending,
        roles=ROLES
    )


# ─── API: создать документ ───────────────────────────────────────────────
@app.route("/api/documents", methods=["POST"])
@login_required
def create_document():
    u   = request.current_user
    d   = request.json or {}
    doc_type = d.get("doc_type")
    if doc_type not in APPROVAL_CHAIN:
        return jsonify({"error": "Неизвестный тип документа"}), 400
    title       = (d.get("title") or "").strip()
    description = (d.get("description") or "").strip()
    priority    = d.get("priority", "normal")
    if priority not in ("low","normal","high","urgent"):
        priority = "normal"
    if not title:
        return jsonify({"error": "Укажите название документа"}), 400
    chain = APPROVAL_CHAIN[doc_type]
    first_step = chain[0]
    with get_db() as db:
        doc_num = _gen_doc_number(db, doc_type)
        cur = db.execute(
            """INSERT INTO documents
               (doc_number,doc_type,title,description,priority,status,
                current_step,current_role,created_by_id,created_by_name,
                item_id,item_inv,department,amount,deadline,employee_id,employee_name)
               VALUES (?,?,?,?,?,'pending',?,?,?,?,?,?,?,?,?,?,?)""",
            (doc_num, doc_type, title, description, priority,
             first_step["step"], first_step["role"],
             u["id"], u["name"],
             d.get("item_id"), d.get("item_inv"),
             d.get("department", u.get("department","")),
             d.get("amount"), d.get("deadline"),
             d.get("employee_id"), d.get("employee_name"))
        )
        doc_id = cur.lastrowid
        # Создать записи согласования для всех шагов
        for step in chain:
            db.execute(
                """INSERT INTO doc_approvals (doc_id,step,role,role_label)
                   VALUES (?,?,?,?)""",
                (doc_id, step["step"], step["role"], step["label"])
            )
        # Автокомментарий
        db.execute(
            "INSERT INTO doc_comments (doc_id,user_id,user_name,user_role,text) VALUES (?,?,?,?,?)",
            (doc_id, u["id"], u["name"], u["role"], f"Документ создан. Ожидает согласования: {first_step['label']}")
        )
        # Уведомление первого согласующего
        msg = f"📄 <b>Новый документ</b>\n{title}\nОжидает вашего согласования."
        notify_role(first_step["role"], msg)
    return jsonify({"ok": True, "doc_id": doc_id, "doc_number": doc_num})


# ─── API: список документов ─────────────────────────────────────────────
@app.route("/api/documents")
@login_required
def list_documents():
    u    = request.current_user
    role = u["role"]
    status_f = request.args.get("status","")
    type_f   = request.args.get("doc_type","")
    with get_db() as db:
        # Расширенный список ролей, которые видят всё
        ADMIN_ROLES = ("superadmin","aho","deputy","director","accountant","auditor","viewer")
        
        if role in ADMIN_ROLES:
            where = "1=1"
            params = []
        else:
            # Обычный сотрудник видит свои документы и те, где он является объектом
            where = "(created_by_id=? OR employee_id=?)"
            params = [u["id"], u["id"]]
        if status_f:
            where += " AND status=?"; params.append(status_f)
        if type_f:
            where += " AND doc_type=?"; params.append(type_f)
        docs = db.execute(
            f"""SELECT d.*,
                (SELECT COUNT(*) FROM doc_comments WHERE doc_id=d.id) as comment_count,
                (SELECT COUNT(*) FROM doc_approvals WHERE doc_id=d.id AND action='approved') as approved_steps,
                (SELECT COUNT(*) FROM doc_approvals WHERE doc_id=d.id) as total_steps
            FROM documents d WHERE {where} ORDER BY d.created_at DESC LIMIT 100""",
            params
        ).fetchall()
    return jsonify([dict(r) for r in docs])


# ─── API: детали документа ──────────────────────────────────────────────
@app.route("/api/documents/<int:did>")
@login_required
def get_document(did):
    u = request.current_user
    with get_db() as db:
        doc = db.execute("SELECT * FROM documents WHERE id=?", (did,)).fetchone()
        if not doc:
            return jsonify({"error": "Не найдено"}), 404
        # Проверка доступа
        if u["role"] not in ("superadmin",) and \
           doc["created_by_id"] != u["id"] and \
           doc["employee_id"] != u["id"] and \
           u["role"] not in ("aho","deputy","director","accountant"):
            return jsonify({"error": "Нет доступа"}), 403
        approvals = db.execute(
            "SELECT * FROM doc_approvals WHERE doc_id=? ORDER BY step",
            (did,)
        ).fetchall()
        comments = db.execute(
            "SELECT * FROM doc_comments WHERE doc_id=? ORDER BY created_at",
            (did,)
        ).fetchall()
    return jsonify({
        "doc": dict(doc),
        "approvals": [dict(a) for a in approvals],
        "comments": [dict(c) for c in comments],
        "chain": APPROVAL_CHAIN.get(doc["doc_type"], []),
    })


# ─── API: согласовать / отклонить ────────────────────────────────────────
@app.route("/api/documents/<int:did>/approve", methods=["POST"])
@login_required
def approve_document(did):
    u      = request.current_user
    d      = request.json or {}
    action  = d.get("action")  # "approved" | "rejected"
    comment = (d.get("comment") or "").strip()[:1000]
    if action not in ("approved", "rejected"):
        return jsonify({"error": "Неверное действие"}), 400
    with get_db() as db:
        doc = db.execute("SELECT * FROM documents WHERE id=?", (did,)).fetchone()
        if not doc:
            return jsonify({"error": "Документ не найден"}), 404
        if doc["status"] != "pending":
            return jsonify({"error": "Документ уже обработан"}), 400
        # Superadmin НЕ должен подписывать за другие роли
        if doc["current_role"] != u["role"]:
            return jsonify({"error": f"Это действие только для роли: {doc['current_role']}"}), 403
        # Если это Superadmin, но роль совпадает - ок. Но обычно Superadmin не в цепочке.
        # Обновить текущий шаг согласования
        sig_path = None
        if action == "approved" and d.get("signature"):
            sig_path = _save_signature(d["signature"], f"doc_{did}_step_{doc['current_step']}")

        db.execute(
            """UPDATE doc_approvals SET action=?,approver_id=?,approver_name=?,
               comment=?,signature=?,acted_at=CURRENT_TIMESTAMP
               WHERE doc_id=? AND step=?""",
            (action, u["id"], u["name"], comment, sig_path, did, doc["current_step"])
        )
        # Добавить комментарий
        action_label = "✅ Согласовал" if action=="approved" else "❌ Отклонил"
        auto_comment = f"{action_label}: {u['name']} ({ROLES.get(u['role'],{}).get('label',u['role'])})"
        if comment:
            auto_comment += f"\nКомментарий: {comment}"
        db.execute(
            "INSERT INTO doc_comments (doc_id,user_id,user_name,user_role,text) VALUES (?,?,?,?,?)",
            (did, u["id"], u["name"], u["role"], auto_comment)
        )
        if action == "rejected":
            db.execute("UPDATE documents SET status='rejected',updated_at=CURRENT_TIMESTAMP WHERE id=?", (did,))
            new_status = "rejected"
        else:
            # Найти следующий шаг
            nxt = _next_doc_step(doc["doc_type"], doc["current_step"])
            if nxt:
                # Следующий согласующий
                db.execute(
                    "UPDATE documents SET current_step=?,current_role=?,updated_at=CURRENT_TIMESTAMP WHERE id=?",
                    (nxt["step"], nxt["role"], did)
                )
                new_status = "pending"
                db.execute(
                    "INSERT INTO doc_comments (doc_id,user_id,user_name,user_role,text) VALUES (?,?,?,?,?)",
                    (did, u["id"], u["name"], u["role"], f"Передано на согласование: {nxt['label']}")
                )
                notify_role(nxt["role"], f"📄 <b>Документ на подпись</b>\n{doc['title']}\nПередано вам на этап: {nxt['label']}")
            else:
                # Все шаги пройдены — документ утверждён
                db.execute(
                    "UPDATE documents SET status='approved',current_role=NULL,updated_at=CURRENT_TIMESTAMP WHERE id=?",
                    (did,)
                )
                new_status = "approved"
                db.execute(
                    "INSERT INTO doc_comments (doc_id,user_id,user_name,user_role,text) VALUES (?,?,?,?,?)",
                    (did, u["id"], u["name"], u["role"], "🎉 Документ полностью согласован и утверждён!")
                )
                notify_user(doc["created_by_id"], f"✅ <b>Документ утверждён!</b>\n{doc['title']}\nВсе этапы согласования пройдены.")
        if action == "rejected":
             notify_user(doc["created_by_id"], f"❌ <b>Документ отклонен</b>\n{doc['title']}\nПричина: {comment or 'не указана'}")
    return jsonify({"ok": True, "new_status": new_status})


# ─── API: добавить комментарий ───────────────────────────────────────────
@app.route("/api/documents/<int:did>/comments", methods=["POST"])
@login_required
def add_doc_comment(did):
    u    = request.current_user
    text = (request.json or {}).get("text","").strip()
    if not text or len(text) > 2000:
        return jsonify({"error": "Некорректный комментарий"}), 400
    with get_db() as db:
        doc = db.execute("SELECT id FROM documents WHERE id=?", (did,)).fetchone()
        if not doc:
            return jsonify({"error": "Не найдено"}), 404
        db.execute(
            "INSERT INTO doc_comments (doc_id,user_id,user_name,user_role,text) VALUES (?,?,?,?,?)",
            (did, u["id"], u["name"], u["role"], text)
        )
    return jsonify({"ok": True})


# ─── API: пометить как распечатано (бухгалтер) ───────────────────────────
@app.route("/api/documents/<int:did>/print", methods=["POST"])
@login_required
def mark_printed(did):
    u = request.current_user
    if u["role"] not in ("accountant","superadmin"):
        return jsonify({"error": "Только бухгалтер"}), 403
    with get_db() as db:
        doc = db.execute("SELECT * FROM documents WHERE id=?", (did,)).fetchone()
        if not doc:
            return jsonify({"error": "Не найдено"}), 404
        if doc["status"] != "approved":
            return jsonify({"error": "Документ ещё не утверждён"}), 400
        db.execute(
            "UPDATE documents SET status='printed',closed_at=CURRENT_TIMESTAMP,updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (did,)
        )
        db.execute(
            "INSERT INTO doc_comments (doc_id,user_id,user_name,user_role,text) VALUES (?,?,?,?,?)",
            (did, u["id"], u["name"], u["role"], f"🖨️ Документ распечатан и закрыт: {u['name']}")
        )
    return jsonify({"ok": True})


# ─── API: статистика документов ──────────────────────────────────────────
@app.route("/api/documents/stats")
@login_required
def doc_stats():
    u = request.current_user
    with get_db() as db:
        total    = db.execute("SELECT COUNT(*) FROM documents").fetchone()[0]
        pending  = db.execute("SELECT COUNT(*) FROM documents WHERE status='pending'").fetchone()[0]
        approved = db.execute("SELECT COUNT(*) FROM documents WHERE status='approved'").fetchone()[0]
        rejected = db.execute("SELECT COUNT(*) FROM documents WHERE status='rejected'").fetchone()[0]
        printed  = db.execute("SELECT COUNT(*) FROM documents WHERE status='printed'").fetchone()[0]
        my_pending = db.execute(
            "SELECT COUNT(*) FROM documents WHERE status='pending' AND current_role=?",
            (u["role"],)
        ).fetchone()[0]
        # Типы
        by_type = db.execute(
            "SELECT doc_type,COUNT(*) as cnt FROM documents GROUP BY doc_type"
        ).fetchall()
    return jsonify({
        "total": total, "pending": pending, "approved": approved,
        "rejected": rejected, "printed": printed,
        "my_pending": my_pending,
        "by_type": [dict(r) for r in by_type]
    })



# ═══════════════════════════════════════════════════════════════════════════
#  NEW FEATURES MODULE — v6
# ═══════════════════════════════════════════════════════════════════════════

# ─── TELEGRAM NOTIFICATIONS ───────────────────────────────────────────────
def send_telegram(chat_id, text):
    """Отправить уведомление в Telegram. chat_id — числовой ID чата."""
    import urllib.request as _ur
    tok = os.environ.get("TELEGRAM_BOT_TOKEN", "")
    if not tok or not chat_id:
        return False
    try:
        payload = json.dumps({"chat_id": chat_id, "text": text, "parse_mode": "HTML"}).encode()
        req = _ur.Request(
            f"https://api.telegram.org/bot{tok}/sendMessage",
            data=payload, headers={"Content-Type": "application/json"}, method="POST"
        )
        _ur.urlopen(req, timeout=4)
        return True
    except Exception as e:
        app.logger.warning(f"Telegram send failed: {e}")
        return False

def notify_user(user_id, text):
    """Отправить уведомление пользователю (Telegram если есть)."""
    def _do():
        try:
            with get_db() as db:
                u = db.execute("SELECT telegram_chat_id FROM users WHERE id=?", (user_id,)).fetchone()
                if u and u["telegram_chat_id"]:
                    send_telegram(u["telegram_chat_id"], text)
        except Exception as e:
            app.logger.warning(f"notify_user error: {e}")
    import threading
    threading.Thread(target=_do, daemon=True).start()

def notify_role(role, text):
    """Отправить уведомление всем пользователям с определенной ролью."""
    try:
        with get_db() as db:
            users = db.execute("SELECT id FROM users WHERE role=? AND active=1", (role,)).fetchall()
            for u in users:
                notify_user(u["id"], text)
    except Exception as e:
        app.logger.warning(f"notify_role error: {role}, {e}")

@app.route("/api/telegram/verify", methods=["POST"])
@login_required
def verify_telegram():
    """Пользователь вводит свой Telegram chat_id и мы отправляем тестовое сообщение."""
    u = request.current_user
    chat_id = (request.json or {}).get("chat_id", "")
    if not chat_id:
        return jsonify({"error": "Укажите chat_id"}), 400
    ok = send_telegram(chat_id, f"✅ <b>Tracko</b>\nПривет, {u['name']}! Уведомления подключены.")
    if ok:
        with get_db() as db:
            db.execute("UPDATE users SET telegram_chat_id=? WHERE id=?", (str(chat_id), u["id"]))
        return jsonify({"ok": True})
    return jsonify({"error": "Не удалось отправить. Проверьте chat_id и бот-токен"}), 400

@app.route("/api/telegram/disconnect", methods=["POST"])
@login_required
def disconnect_telegram():
    u = request.current_user
    with get_db() as db:
        db.execute("UPDATE users SET telegram_chat_id=NULL WHERE id=?", (u["id"],))
    return jsonify({"ok": True})


# ─── 2FA / TOTP ────────────────────────────────────────────────────────────
@app.route("/api/2fa/setup", methods=["POST"])
@login_required
def totp_setup():
    """Сгенерировать TOTP секрет и QR для Google Authenticator."""
    u = request.current_user
    if u["role"] not in ("superadmin", "director", "accountant", "deputy"):
        return jsonify({"error": "2FA доступна только для администраторов"}), 403
    secret = pyotp.random_base32()
    totp = pyotp.TOTP(secret)
    uri = totp.provisioning_uri(name=u["email"], issuer_name="Tracko")
    # Generate QR PNG as base64
    import io as _io
    qr_img = qrcode.make(uri)
    buf = _io.BytesIO(); qr_img.save(buf, "PNG"); buf.seek(0)
    import base64 as _b64
    qr_b64 = _b64.b64encode(buf.read()).decode()
    # Save secret (not enabled yet — user must confirm)
    with get_db() as db:
        db.execute("UPDATE users SET totp_secret=?,totp_enabled=0 WHERE id=?", (secret, u["id"]))
    return jsonify({"ok": True, "secret": secret, "qr": f"data:image/png;base64,{qr_b64}", "uri": uri})

@app.route("/api/2fa/confirm", methods=["POST"])
@login_required
def totp_confirm():
    """Подтвердить активацию 2FA кодом из приложения."""
    u = request.current_user
    code = (request.json or {}).get("code", "")
    with get_db() as db:
        row = db.execute("SELECT totp_secret FROM users WHERE id=?", (u["id"],)).fetchone()
        if not row or not row["totp_secret"]:
            return jsonify({"error": "Сначала настройте 2FA"}), 400
        totp = pyotp.TOTP(row["totp_secret"])
        if not totp.verify(code, valid_window=1):
            return jsonify({"error": "Неверный код"}), 400
        db.execute("UPDATE users SET totp_enabled=1 WHERE id=?", (u["id"],))
    return jsonify({"ok": True})

@app.route("/api/2fa/disable", methods=["POST"])
@login_required
def totp_disable():
    """Отключить 2FA."""
    u = request.current_user
    code = (request.json or {}).get("code", "")
    with get_db() as db:
        row = db.execute("SELECT totp_secret,totp_enabled FROM users WHERE id=?", (u["id"],)).fetchone()
        if row and row["totp_enabled"]:
            totp = pyotp.TOTP(row["totp_secret"])
            if not totp.verify(code, valid_window=1):
                return jsonify({"error": "Неверный код для отключения"}), 400
        db.execute("UPDATE users SET totp_enabled=0,totp_secret=NULL WHERE id=?", (u["id"],))
    return jsonify({"ok": True})


# ─── PLANNER INVENTORY SESSIONS ────────────────────────────────────────────
@app.route("/api/inventory/sessions", methods=["GET"])
@roles_required("superadmin", "aho", "auditor")
def list_inv_sessions():
    with get_db() as db:
        rows = db.execute(
            "SELECT * FROM inventory_sessions ORDER BY created_at DESC LIMIT 50"
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/inventory/sessions", methods=["POST"])
@roles_required("superadmin", "aho")
def create_inv_session():
    u = request.current_user
    d = request.json or {}
    title = (d.get("title") or "").strip()
    if not title:
        return jsonify({"error": "Укажите название"}), 400
    dept = d.get("department", "")
    with get_db() as db:
        # Collect items to check
        if dept:
            items = db.execute(
                "SELECT i.id FROM items i JOIN users u ON u.name=i.employee WHERE u.department=?", (dept,)
            ).fetchall()
        else:
            items = db.execute("SELECT id FROM items WHERE status='Занято'").fetchall()
        cur = db.execute(
            "INSERT INTO inventory_sessions (title,created_by_id,created_by_name,department,total_items) VALUES (?,?,?,?,?)",
            (title, u["id"], u["name"], dept, len(items))
        )
        sid = cur.lastrowid
        for item in items:
            db.execute(
                "INSERT INTO inventory_checks (session_id,item_id) VALUES (?,?)",
                (sid, item["id"])
            )
    return jsonify({"ok": True, "session_id": sid, "total_items": len(items)})

@app.route("/api/inventory/sessions/<int:sid>", methods=["GET"])
@login_required
def get_inv_session(sid):
    with get_db() as db:
        session = db.execute("SELECT * FROM inventory_sessions WHERE id=?", (sid,)).fetchone()
        if not session:
            return jsonify({"error": "Не найдено"}), 404
        checks = db.execute(
            """SELECT c.*,i.inv_num,i.category,i.model,i.room,i.employee
               FROM inventory_checks c
               JOIN items i ON c.item_id=i.id
               WHERE c.session_id=? ORDER BY i.room,i.category""",
            (sid,)
        ).fetchall()
    return jsonify({"session": dict(session), "checks": [dict(r) for r in checks]})

@app.route("/api/inventory/check/<int:cid>", methods=["POST"])
@login_required
def submit_inv_check(cid):
    """Сотрудник/АХО подтверждает наличие актива."""
    u = request.current_user
    d = request.json or {}
    status = d.get("status", "found")  # found | not_found | damaged
    note   = (d.get("note") or "").strip()[:500]
    with get_db() as db:
        check = db.execute("SELECT * FROM inventory_checks WHERE id=?", (cid,)).fetchone()
        if not check:
            return jsonify({"error": "Не найдено"}), 404
        db.execute(
            """UPDATE inventory_checks SET status=?,checked_by_id=?,checked_by_name=?,
               note=?,checked_at=CURRENT_TIMESTAMP WHERE id=?""",
            (status, u["id"], u["name"], note, cid)
        )
        # Update session progress
        session_id = check["session_id"]
        checked = db.execute(
            "SELECT COUNT(*) FROM inventory_checks WHERE session_id=? AND status!='pending'",
            (session_id,)
        ).fetchone()[0]
        total = db.execute(
            "SELECT total_items FROM inventory_sessions WHERE id=?", (session_id,)
        ).fetchone()[0]
        db.execute(
            "UPDATE inventory_sessions SET checked_items=? WHERE id=?",
            (checked, session_id)
        )
        if checked >= total:
            db.execute(
                "UPDATE inventory_sessions SET status='completed',completed_at=CURRENT_TIMESTAMP WHERE id=?",
                (session_id,)
            )
    return jsonify({"ok": True, "progress": f"{checked}/{total}"})


# ─── EQUIPMENT TEMPLATES ──────────────────────────────────────────────────
@app.route("/api/equipment-templates", methods=["GET"])
@login_required
def list_eq_templates():
    with get_db() as db:
        rows = db.execute("SELECT * FROM equipment_templates ORDER BY name").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/equipment-templates", methods=["POST"])
@roles_required("superadmin", "aho")
def create_eq_template():
    u = request.current_user
    d = request.json or {}
    name = (d.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Укажите название шаблона"}), 400
    items = d.get("items", [])  # [{"category":"Ноутбук","model":"..."}, ...]
    with get_db() as db:
        cur = db.execute(
            "INSERT INTO equipment_templates (name,description,items_json,created_by) VALUES (?,?,?,?)",
            (name, d.get("description",""), json.dumps(items, ensure_ascii=False), u["name"])
        )
    return jsonify({"ok": True, "id": cur.lastrowid})

@app.route("/api/equipment-templates/<int:tid>/apply", methods=["POST"])
@roles_required("superadmin", "aho", "hr")
def apply_eq_template(tid):
    """Применить шаблон к новому сотруднику — создать все активы."""
    u = request.current_user
    d = request.json or {}
    employee_name = (d.get("employee_name") or "").strip()
    employee_id   = d.get("employee_id")
    room          = (d.get("room") or "").strip()
    place         = (d.get("place") or "").strip()
    if not employee_name:
        return jsonify({"error": "Укажите сотрудника"}), 400
    with get_db() as db:
        tmpl = db.execute("SELECT * FROM equipment_templates WHERE id=?", (tid,)).fetchone()
        if not tmpl:
            return jsonify({"error": "Шаблон не найден"}), 404
        items_def = json.loads(tmpl["items_json"] or "[]")
        created_ids = []
        for item_def in items_def:
            cat   = item_def.get("category", "Другое")
            model = item_def.get("model", "")
            inv   = next_inv(cat)
            cur = db.execute(
                """INSERT INTO items (inv_num,category,model,room,place,employee,employee_id,
                   status,condition,check_date,notes)
                   VALUES (?,?,?,?,?,?,?,'Занято','Хорошее',?,?)""",
                (inv, cat, model, room, place, employee_name, employee_id,
                 date.today().isoformat(), f"Шаблон: {tmpl['name']}")
            )
            log_h(db, cur.lastrowid, f"Создан по шаблону: {tmpl['name']}", uid=u["id"], uname=u["name"])
            created_ids.append({"id": cur.lastrowid, "inv_num": inv, "category": cat})
    return jsonify({"ok": True, "created": created_ids, "count": len(created_ids)})


# ─── API KEYS ──────────────────────────────────────────────────────────────
import hashlib as _hashlib
import secrets as _sec_mod

@app.route("/api/api-keys", methods=["GET"])
@roles_required("superadmin")
def list_api_keys():
    with get_db() as db:
        rows = db.execute(
            "SELECT id,name,scopes,last_used,expires_at,active,created_at FROM api_keys ORDER BY created_at DESC"
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/api-keys", methods=["POST"])
@roles_required("superadmin")
def create_api_key():
    u = request.current_user
    d = request.json or {}
    name = (d.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Укажите название ключа"}), 400
    raw_key = "trk_" + _sec_mod.token_hex(32)
    key_hash = _hashlib.sha256(raw_key.encode()).hexdigest()
    with get_db() as db:
        cur = db.execute(
            "INSERT INTO api_keys (name,key_hash,scopes,user_id,expires_at) VALUES (?,?,?,?,?)",
            (name, key_hash, d.get("scopes","read"), u["id"], d.get("expires_at"))
        )
    return jsonify({"ok": True, "id": cur.lastrowid, "key": raw_key,
                    "warning": "Сохраните ключ — он показывается только один раз!"})

@app.route("/api/api-keys/<int:kid>", methods=["DELETE"])
@roles_required("superadmin")
def revoke_api_key(kid):
    with get_db() as db:
        db.execute("UPDATE api_keys SET active=0 WHERE id=?", (kid,))
    return jsonify({"ok": True})


# ─── ONBOARDING COMPLETE ──────────────────────────────────────────────────
@app.route("/api/onboarding/complete", methods=["POST"])
@login_required
def complete_onboarding():
    u = request.current_user
    with get_db() as db:
        db.execute("UPDATE users SET onboarding_done=1 WHERE id=?", (u["id"],))
    return jsonify({"ok": True})


# ─── DEPRECIATION / ASSET VALUE ──────────────────────────────────────────
@app.route("/api/items/<int:iid>/valuation")
@login_required
def item_valuation(iid):
    """Рассчитать остаточную стоимость актива (линейная амортизация)."""
    with get_db() as db:
        item = db.execute("SELECT * FROM items WHERE id=?", (iid,)).fetchone()
    if not item:
        return jsonify({"error": "Не найдено"}), 404
    item = dict(item)
    if not item.get("purchase_price") or not item.get("purchase_date"):
        return jsonify({"purchase_price": None, "residual_value": None,
                        "depreciation_pct": None, "message": "Нет данных о стоимости"})
    # Default useful life by category (years)
    useful_life = {
        "Ноутбук": 3, "Монитор": 5, "Кресло": 7, "Стол": 10,
        "Принтер": 4, "Телефон": 2, "Клавиатура": 3, "Мышь": 2,
    }.get(item.get("category",""), 5)
    from datetime import date as _d
    today = _d.today()
    try:
        purchased = _d.fromisoformat(item["purchase_date"])
    except:
        return jsonify({"error": "Неверный формат даты покупки"}), 400
    years_used = (today - purchased).days / 365.25
    depreciation_pct = min(100.0, (years_used / useful_life) * 100)
    residual_value = round(item["purchase_price"] * (1 - depreciation_pct / 100), 2)
    return jsonify({
        "purchase_price": item["purchase_price"],
        "purchase_date": item["purchase_date"],
        "years_used": round(years_used, 1),
        "useful_life_years": useful_life,
        "depreciation_pct": round(depreciation_pct, 1),
        "residual_value": max(0, residual_value),
        "category": item.get("category")
    })


# ─── GLOBAL SEARCH (enhanced) ────────────────────────────────────────────
@app.route("/api/search/unified")
@login_required
def unified_search():
    """Единый поиск по активам, документам, сотрудникам."""
    u   = request.current_user
    q   = (request.args.get("q") or "").strip()[:100]
    if len(q) < 2:
        return jsonify({"items":[],"documents":[],"employees":[]})
    pat = f"%{q}%"
    with get_db() as db:
        # Items
        if u["role"] in ("employee", "viewer"):
            items = db.execute(
                """SELECT id,inv_num,category,model,room,employee,status FROM items
                   WHERE (inv_num LIKE ? OR model LIKE ? OR serial_num LIKE ? OR employee LIKE ?)
                   AND (employee_id=? OR employee=?) LIMIT 8""",
                (pat,pat,pat,pat,u["id"],u["name"])
            ).fetchall()
        else:
            items = db.execute(
                """SELECT id,inv_num,category,model,room,employee,status FROM items
                   WHERE inv_num LIKE ? OR model LIKE ? OR serial_num LIKE ? OR employee LIKE ?
                   LIMIT 8""",
                (pat,pat,pat,pat)
            ).fetchall()
        # Documents
        is_admin = u["role"] in ('superadmin','aho','deputy','director','accountant','auditor')
        docs = db.execute(
            f"""SELECT id,doc_number,doc_type,title,status,created_by_name FROM documents
               WHERE (title LIKE ? OR doc_number LIKE ? OR description LIKE ?)
               {"AND (created_by_id=? OR ?)" if not is_admin else ""}
               LIMIT 5""",
            (pat,pat,pat,u["id"], "1=1") if not is_admin else (pat,pat,pat)
        ).fetchall()
        # Employees
        if u["role"] not in ("employee", "viewer"):
            emps = db.execute(
                """SELECT id,name,email,role,department FROM users
                   WHERE (name LIKE ? OR email LIKE ?) AND active=1 LIMIT 5""",
                (pat,pat)
            ).fetchall()
        else:
            emps = []
    return jsonify({
        "items":     [dict(r) for r in items],
        "documents": [dict(r) for r in docs],
        "employees": [dict(r) for r in emps],
    })


# ─── RATE LIMITING (enhanced) ────────────────────────────────────────────
_RATE_STORE: dict = {}  # ip -> {"count":N, "reset":timestamp}
RATE_LIMIT_DEFAULT = 300  # requests per minute for API
RATE_LIMIT_LOGIN   = 5    # login attempts per 5 min

def check_global_rate_limit(ip: str, limit: int = RATE_LIMIT_DEFAULT, window: int = 60) -> bool:
    """True = allowed, False = rate limited."""
    now = time.time()
    key = f"global:{ip}"
    if key not in _RATE_STORE or now > _RATE_STORE[key]["reset"]:
        _RATE_STORE[key] = {"count": 1, "reset": now + window}
        return True
    _RATE_STORE[key]["count"] += 1
    return _RATE_STORE[key]["count"] <= limit

@app.before_request
def global_rate_limit():
    """Apply global rate limit to all API endpoints."""
    if request.path.startswith("/api/") and request.path not in ("/api/auth/login",):
        ip = request.remote_addr or "unknown"
        if not check_global_rate_limit(ip):
            return jsonify({"error": "Слишком много запросов. Подождите минуту."}), 429


# ─── SETTINGS PAGE ────────────────────────────────────────────────────────
@app.route("/settings")
@login_required
def settings_page():
    u = request.current_user
    with get_db() as db:
        # Check 2FA status
        row = db.execute("SELECT totp_enabled,telegram_chat_id FROM users WHERE id=?", (u["id"],)).fetchone()
    return render_template("settings.html",
        user=u, current_user=u,
        role_info=ROLES.get(u["role"], {}),
        totp_enabled=bool(row["totp_enabled"]) if row else False,
        telegram_connected=bool(row["telegram_chat_id"]) if row else False,
        roles=ROLES
    )
# ─── KNOWLEDGE BASE ───────────────────────────────────────────────────────────
@app.route("/docs")
@login_required
def docs_page():
    u = request.current_user
    return render_template("docs.html",
        user=u, current_user=u,
        role_info=ROLES.get(u["role"], {}),
        roles=ROLES)

# ─── FINANCIALS ────────────────────────────────────────────────────────────────
@app.route("/api/financials")
@roles_required("superadmin","aho","director","deputy","accountant","auditor")
def financials():
    with get_db() as db:
        total = db.execute("SELECT COALESCE(SUM(purchase_price),0) FROM items WHERE purchase_price IS NOT NULL").fetchone()[0]
        by_cat = db.execute("SELECT category,COUNT(*) cnt,COALESCE(SUM(purchase_price),0) total_val FROM items GROUP BY category ORDER BY total_val DESC").fetchall()
    return jsonify({"total_purchase_value":round(total,2),"by_category":[dict(r) for r in by_cat]})

# ─── ACTIVE USERS ─────────────────────────────────────────────────────────────
@app.route("/api/users/active")
@login_required
def users_active():
    with get_db() as db:
        rows = db.execute("SELECT id,name,email,role,department FROM users WHERE active=1 ORDER BY name").fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/notifications")
@login_required
def get_notifications():
    """Real-time notifications for nav badge."""
    u = request.current_user
    with get_db() as db:
        notifs = []
        if u["role"] in ("superadmin","aho","deputy","director","accountant"):
            # Pending documents awaiting this role
            pending_docs = db.execute(
                "SELECT COUNT(*) FROM documents WHERE status='pending' AND current_role=?",
                (u["role"],)
            ).fetchone()[0]
            if pending_docs > 0:
                notifs.append({"type":"docs","count":pending_docs,"label":f"{pending_docs} документ(а) ждут согласования"})
        if u["role"] in ("superadmin","aho"):
            # Items requiring repair
            repair = db.execute(
                "SELECT COUNT(*) FROM maintenance WHERE status='pending'"
            ).fetchone()[0]
            if repair > 0:
                notifs.append({"type":"repair","count":repair,"label":f"{repair} заявок на ремонт"})
            # Overdue audit (not checked in 180 days)
            overdue = db.execute(
                "SELECT COUNT(*) FROM items WHERE check_date < date('now','-180 days') OR check_date IS NULL"
            ).fetchone()[0]
            if overdue > 0:
                notifs.append({"type":"audit","count":overdue,"label":f"{overdue} активов без проверки 180+ дней"})
        
        # Notifications for Employee
        if u["role"] == "employee":
            # Approved/Rejected documents
            my_docs = db.execute(
                "SELECT COUNT(*) FROM documents WHERE created_by_id=? AND status IN ('approved','rejected') AND updated_at > datetime('now','-7 days')",
                (u["id"],)
            ).fetchone()[0]
            if my_docs > 0:
                notifs.append({"type":"docs","count":my_docs,"label":f"У вас {my_docs} обновленных документа"})
            
            # Pending issuances (confirm receipt)
            pending_iss = db.execute(
                "SELECT COUNT(*) FROM issuances WHERE employee_id=? AND status='pending'",
                (u["id"],)
            ).fetchone()[0]
            if pending_iss > 0:
                notifs.append({"type":"issuance","count":pending_iss,"label":f"Ожидает подтверждения: {pending_iss} выдачи"})

        total = sum(n["count"] for n in notifs)
    return jsonify({"total": total, "items": notifs})



@app.errorhandler(404)
def page_not_found(e):
    return render_template('index.html'), 404

@app.errorhandler(500)
def internal_error(e):
    return jsonify({"error": "Internal Server Error", "message": str(e)}), 500

# ══════════════════════════════════════════════════════════
# MISSING PAGE ROUTES
# ══════════════════════════════════════════════════════════

@app.route("/maintenance")
@login_required
def maintenance_page():
    u = request.current_user
    return render_template("maintenance.html", user=u, current_user=u,
        role_info=ROLES.get(u["role"],{}), roles=ROLES)

@app.route("/requests")
@login_required
def requests_page():
    u = request.current_user
    return render_template("requests.html", user=u, current_user=u,
        role_info=ROLES.get(u["role"],{}), roles=ROLES)

@app.route("/analytics")
@roles_required("superadmin","aho","director","deputy","auditor","accountant")
def analytics_page():
    u = request.current_user
    return render_template("analytics.html", user=u, current_user=u,
        role_info=ROLES.get(u["role"],{}), roles=ROLES)

@app.route("/inventory")
@roles_required("superadmin","aho","auditor")
def inventory_page():
    u = request.current_user
    return render_template("inventory.html", user=u, current_user=u,
        role_info=ROLES.get(u["role"],{}), roles=ROLES)

@app.route("/security")
@roles_required("superadmin")
def security_page():
    u = request.current_user
    with get_db() as db:
        try:
            logs = db.execute(
                "SELECT l.*,COALESCE(u2.name,l.email) as name "
                "FROM login_log l LEFT JOIN users u2 ON l.user_id=u2.id "
                "ORDER BY l.ts DESC LIMIT 300"
            ).fetchall()
            logs = [dict(r) for r in logs]
        except Exception: logs = []
    return render_template("security.html", user=u, current_user=u,
        role_info=ROLES.get(u["role"],{}), roles=ROLES, logs=logs)

@app.route("/billing")
@roles_required("superadmin","director","accountant")
def billing_page():
    u = request.current_user
    PLANS_CFG = {
        "starter":    {"label":"Старт",      "price":"$30/мес",  "max_items":50,   "max_users":5},
        "business":   {"label":"Бизнес",     "price":"$80/мес",  "max_items":500,  "max_users":25},
        "enterprise": {"label":"Корпоратив", "price":"$300+/мес","max_items":None, "max_users":None},
    }
    with get_db() as db:
        try:
            sub = db.execute("SELECT * FROM subscriptions ORDER BY id DESC LIMIT 1").fetchone()
            sub = dict(sub) if sub else {"plan":"starter","max_items":50,"max_users":5,"expires_at":None}
            used_items = db.execute("SELECT COUNT(*) FROM items").fetchone()[0]
            used_users = db.execute("SELECT COUNT(*) FROM users WHERE active=1").fetchone()[0]
        except Exception: sub={"plan":"starter"}; used_items=0; used_users=0
    return render_template("billing.html", user=u, current_user=u,
        role_info=ROLES.get(u["role"],{}), roles=ROLES,
        plans=PLANS_CFG, subscription=sub,
        used_items=used_items, used_users=used_users)

# ══════════════════════════════════════════════════════════
# DISMISSAL EXPORT ACT (Excel)
# ══════════════════════════════════════════════════════════
@app.route("/api/dismissals/<int:did>/export")
@roles_required("superadmin","aho","hr","accountant")
def export_dismissal_act(did):
    from openpyxl.styles import Font,Alignment,PatternFill,Border,Side
    from openpyxl.utils import get_column_letter
    with get_db() as db:
        dis = db.execute("SELECT * FROM dismissals WHERE id=?",(did,)).fetchone()
    if not dis: abort(404)
    dis = dict(dis)
    items_l = json.loads(dis.get("items_json") or "[]")
    cmap    = json.loads(dis.get("item_conditions") or "{}")
    commmap = json.loads(dis.get("item_comments") or "{}")
    wb=Workbook(); ws=wb.active; ws.title="Акт"
    blue="1E40AF"; thin=Side(style="thin",color="CBD5E1")
    bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
    ws.merge_cells("A1:G1")
    ws["A1"]="АКТ ПРИЁМА-ПЕРЕДАЧИ МАТЕРИАЛЬНЫХ ЦЕННОСТЕЙ"
    ws["A1"].font=Font(bold=True,size=13,color="FFFFFF")
    ws["A1"].fill=PatternFill("solid",fgColor=blue)
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=30
    for i,(k,v) in enumerate([
        ("Сотрудник:",dis.get("employee_name","")),
        ("Дата:",date.today().strftime("%d.%m.%Y")),
        ("Инициировал:",dis.get("initiated_by_name",""))],3):
        ws[f"A{i}"]=k; ws[f"A{i}"].font=Font(bold=True)
        ws[f"B{i}"]=v; ws.merge_cells(f"B{i}:G{i}")
    hr=7
    for col,h in enumerate(["Инв.№","Категория","Модель","Кабинет","При выдаче","При возврате","Комментарий"],1):
        c=ws.cell(hr,col,h); c.font=Font(bold=True,color="FFFFFF")
        c.fill=PatternFill("solid",fgColor=blue); c.border=bdr
        c.alignment=Alignment(horizontal="center")
    for ri,item in enumerate(items_l,hr+1):
        iid=str(item.get("id",""))
        cr=cmap.get(iid,"—")
        fg="FEE2E2" if cr=="Утеряно" else ("F8FAFC" if ri%2==0 else "FFFFFF")
        for col,val in enumerate([item.get("inv_num",""),item.get("category",""),
            item.get("model",""),item.get("room",""),"Хорошее",cr,commmap.get(iid,"")],1):
            c=ws.cell(ri,col,val); c.border=bdr
            c.fill=PatternFill("solid",fgColor=fg)
    sr=hr+len(items_l)+2
    ws[f"A{sr}"]="Подпись АХО:"; ws[f"A{sr}"].font=Font(bold=True)
    ws[f"C{sr}"]="____________________"
    ws[f"E{sr}"]="Дата:"; ws[f"F{sr}"]=date.today().strftime("%d.%m.%Y")
    for i,w in enumerate([12,14,22,16,16,16,28],1):
        ws.column_dimensions[get_column_letter(i)].width=w
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    fn=f"Акт_{(dis.get('employee_name') or 'emp').replace(' ','_')}_{date.today()}.xlsx"
    return send_file(buf,as_attachment=True,download_name=fn,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ══════════════════════════════════════════════════════════
# PWA
# ══════════════════════════════════════════════════════════
@app.route("/static/manifest.json")
def pwa_manifest():
    return jsonify({"name":"Tracko — Инвентаризация","short_name":"Tracko",
        "start_url":"/","display":"standalone",
        "background_color":"#000","theme_color":"#007AFF",
        "icons":[{"src":"/api/icon/192","sizes":"192x192","type":"image/png"},
                 {"src":"/api/icon/512","sizes":"512x512","type":"image/png"}]
    }),200,{"Content-Type":"application/manifest+json"}

@app.route("/api/icon/<int:size>")
def app_icon(size):
    from PIL import Image,ImageDraw
    sz=min(max(size,48),512)
    img=Image.new("RGB",(sz,sz),"#007AFF")
    draw=ImageDraw.Draw(img)
    p,bh,sw=sz//6,sz//8,sz//8
    draw.rectangle([p,p,sz-p,p+bh],fill="white")
    draw.rectangle([sz//2-sw//2,p,sz//2+sw//2,sz-p],fill="white")
    buf=io.BytesIO(); img.save(buf,"PNG"); buf.seek(0)
    return send_file(buf,mimetype="image/png")


if __name__ == "__main__":
    init_db()
    migrate_db()
    app.run(debug=False, host=os.environ.get('HOST', '0.0.0.0'), port=int(os.environ.get('PORT', 5000)), threaded=True)
